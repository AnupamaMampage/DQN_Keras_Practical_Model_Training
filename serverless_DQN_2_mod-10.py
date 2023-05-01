import math
import os

import openpyxl
from pandas.tests.io.excel.test_xlrd import xlrd
from sys import maxsize
# os.environ['CUDA_VISIBLE_DEVICES'] = '-1'
import random
import requests
import gym
from gym.spaces import Discrete, Box
import numpy as np
from collections import deque
from tensorflow.python import keras
from tensorflow.keras.models import Model, load_model
from tensorflow.keras.layers import Input, Dense
from tensorflow.keras.optimizers import Adam, RMSprop
# from keras.optimizers import Adam, RMSprop
from kubernetes import client, config, watch
from datetime import datetime
import time
import json
import threading
from collections import deque
from queue import Queue
import xlwt
from tensorflow.python.keras.layers import LSTM
from xlwt import Workbook
from tensorflow.keras.callbacks import TensorBoard
# import keras.backend.tensorflow_backend as backend
import tensorflow as tf

step_count_total = 0

MODEL_NAME = "Serverless_Scheduling"
PATH = '/home/ubuntu/work/logs/'
scheduler_name = "customscheduler"
PROMETHEUS = 'http://45.113.232.196:31495/'

NODE_IP = ["45.113.234.2:9100", "45.113.232.27:9100", "45.113.235.56:9100", "45.113.233.249:9100",
           "45.113.232.251:9100", "45.113.232.173:9100", "45.113.232.33:9100", "115.146.93.183:9100",
           "115.146.92.117:9100", "115.146.93.61:9100"]
NODE_NAME = ["node-3", "node-4", "node-5", "node-6", "node-7", "node-8", "node-9", "node-10",
             "node-11", "node-12"]
NODE_MAP_NAME_IP = {"node-3": "45.113.234.2:9100", "node-4": "45.113.232.27:9100", "node-5": "45.113.235.56:9100",
                    "node-6": "45.113.233.249:9100", "node-7": "45.113.232.251:9100", "node-8": "45.113.232.173:9100",
                    "node-9": "45.113.232.33:9100", "node-10": "115.146.93.183:9100",
                    "node-11": "115.146.92.117:9100", "node-12": "115.146.93.61:9100"}
NODE_MAP_IP_NAME = {"45.113.234.2:9100": "node-3", "45.113.232.27:9100": "node-4", "45.113.235.56:9100": "node-5",
                    "45.113.233.249:9100": "node-6", "45.113.232.251:9100": "node-7", "45.113.232.173:9100": "node-8",
                    "45.113.232.33:9100": "node-9", "115.146.93.183:9100": "node-10",
                    "115.146.92.117:9100": "node-11", "115.146.93.61:9100": "node-12"}
NODE_MAP_IP_INDEX = {"45.113.234.2:9100:9100": 0, "45.113.232.27:9100": 1, "45.113.235.56:9100": 2,
                     "45.113.233.249:9100": 3, "45.113.232.251:9100": 4, "45.113.232.173:9100": 5,
                     "45.113.232.33:9100": 6, "115.146.93.183:9100": 7, "115.146.92.117:9100": 8,
                     "115.146.93.61:9100": 9}

NODE_MAP_NAME_INDEX = {"node-3": 0, "node-4": 1, "node-5": 2, "node-6": 3, "node-7": 4, "node-8": 5, "node-9": 6,
                       "node-10": 7,
                       "node-11": 8, "node-12": 9}
NODE_MAP_INDEX_NAME = {0: "node-3", 1: "node-4", 2: "node-5", 3: "node-6", 4: "node-7", 5: "node-8",
                       6: "node-9", 7: "node-10", 8: "node-11", 9: "node-12"}

NODE_HR_RATE = {"node-3": 0.172, "node-4": 0.172, "node-5": 0.172, "node-6": 0.172, "node-7": 0.172, "node-8": 0.172,
                "node-9": 0.086,
                "node-10": 0.086,
                "node-11": 0.344, "node-12": 0.344, "node-13": 0.344, "node-14": 0.344, "node-15": 0.086,
                "node-16": 0.086, "node-17": 0.086, "node-18": 0.086, "node-19": 0.172,
                "node-20": 0.172,
                "node-21": 0.172, "node-22": 0.172}

# 4>t4g.xlarge> 0.1696  2>t4g.large> 0.0848    8>t4g.2xlarge>0.3392
NODE_CPU_CAP = {"node-3": 4, "node-4": 4, "node-5": 4, "node-6": 4, "node-7": 4, "node-8": 4, "node-9": 2,
                "node-10": 2,
                "node-11": 8, "node-12": 8, "node-13": 8, "node-14": 8, "node-15": 2, "node-16": 2, "node-17": 2,
                "node-18": 2, "node-19": 4,
                "node-20": 4,
                "node-21": 4, "node-22": 4}

active_fns = {"float": 0, "load": 0, "matrix": 0, "primary": 0, "dd": 0, "imageprocess": 0, "imageresize": 0,
              "todoadditem": 0, "tododelitem": 0, "todogitem": 0, "todolistitem": 0, "todoupitem": 0, "uploadimage": 0,
              "uploadvideo": 0, "video": 0}
checked_fns = {"float": False, "load": False, "matrix": False, "primary": False, "dd": False, "imageprocess": False,
               "imageresize": False, "todoadditem": False, "tododelitem": False, "todogitem": False,
               "todolistitem": False, "todoupitem": False, "uploadimage": False, "uploadvideo": False, "video": False}

POD_NW_BW_USAGE = {"float": 0, "matrix": 0, "load": 50, "dd": 0, "test": 0}
# POD_NW_BW_USAGE = {"float": 0, "matrix": 0, "load": 50, "dd": 0}
# POD_NW_BW_USAGE = {"float": 0, "matrix": 0, "load": 50, "dd": 0}
FN_ARRAY = ["float", "load", "matrix", "primary", "dd", "imageprocess", "imageresize", "todoadditem", "tododelitem",
            "todogitem", "todolistitem", "todoupitem", "uploadimage", "uploadvideo", "video"]
FN_NAME_INDEX = {"float": 0, "load": 1, "matrix": 2, "primary": 3, "dd": 4, "imageprocess": 5, "imageresize": 6,
                 "todoadditem": 7, "tododelitem": 8, "todogitem": 9, "todolistitem": 10, "todoupitem": 11,
                 "uploadimage": 12, "uploadvideo": 13, "video": 14}
CPU_REQUESTS = {"float": 0.3, "load": 0.12, "matrix": 0.25, "primary": 0.4, "dd": 0.4, "imageprocess": 0.15,
                "imageresize": 0.11, "todoadditem": 0.05, "tododelitem": 0.08, "todogitem": 0.05, "todolistitem": 0.05,
                "todoupitem": 0.05, "uploadimage": 0.1, "uploadvideo": 0.1, "video": 0.09}

FN_DEADLINE_DICT = {"float": 0.4, "load": 0.1, "matrix": 0.275, "primary": 0.45, "dd": 0.075, "imageprocess": 0.09,
                    "imageresize": 0.120, "todoadditem": 0.05, "tododelitem": 0.13, "todogitem": 0.06,
                    "todolistitem": 0.05,
                    "todoupitem": 0.07, "uploadimage": 0.095, "uploadvideo": 0.070, "video": 0.03}

# CPU_REQUESTS = {"float": 2, "matrix": 3, "load": 0.12, "dd": 0.3, "load2": 0.12, "dd2": 0.3}
FN_DEADLINE = [0.4, 0.1, 0.275, 0.45, 0.075, 0.09, 0.120, 0.05, 0.13, 0.06, 0.05, 0.07, 0.095, 0.07, 0.03]
# MODEL_NAME = "serverless_DQN"

# wb = Workbook()
# sch_data = wb.add_sheet('Scheduled')
# sch_data.write(0, 0, 'Time')
# sch_data.write(0, 1, 'Episode')
# sch_data.write(0, 2, 'Event_No')
# sch_data.write(0, 3, 'Pod')
# sch_data.write(0, 4, 'State')
# sch_data.write(0, 5, 'Action')
# sch_data.write(0, 6, 'Reward')
# sch_data.write(0, 7, 'Next_state')
# sch_data.write(0, 8, 'Done')
# sch_data.write(0, 9, 'Random or NOT')
# sch_data.write(0, 10, 'Pod_created_time')
# sch_data.write(0, 11, 'Fn_latency')
# sch_data.write(0, 12, 'Fn_failure_rate')
# sch_data.write(0, 13, 'CPU_penalty')
# sch_data.write(0, 14, 'Active_node_no')
#
# ep_data = wb.add_sheet('Episodes')
# ep_data.write(0, 0, 'Time')
# ep_data.write(0, 1, 'Episode')
# ep_data.write(0, 2, 'Epsilon')
# ep_data.write(0, 3, 'Ep_reward')

reward_q = Queue()
state_q = Queue()
action_q = Queue()
done_q = Queue()
threadID_q = Queue()
next_state_q = Queue()
reward_dict = {}
thread_finish_q = Queue()
memory = deque(maxlen=2000)
n_step = 5
n_step_buffer = deque(maxlen=n_step)

stop_an_thread = True
training_end = False
no_arrivals = False
episodic_reward = 0
function_latency = 0
total_vm_time_diff = 0
total_vm_cost_diff = 0
cpu_util_penalty = 0
act_node_penalty = 0
pending_pod_list = []
active_node_list = []
node_cpu_requests_dict = {}
node_cpu_requests_ratio_dict = {}
cpu_util_dict = {}
active_nodes = 0
initial_lat = 0
prev_action = 0
vm_up_time_prev = 0
vm_up_time_cost_prev = 0
prev_app = ""
host_p = 0
node_count = 0
act_nodes = 0
act_node_cost = 0
max_arr_rate = 20
max_replicas = 6

epsilon = 1  # exploration rate
epsilon_min = 0.04
epsilon_decay = 0.999
train_start = 100

mode = "random"
rr_node_no = 0
d = {}

gamma_nstep = 0.95


def OurModel(input_shape, action_space):
    X_input = Input(input_shape)

    # 'Dense' is the basic form of a neural network layer
    # Input Layer of state size(4) and Hidden Layer with 512 nodes

    # only FC layers
    X = Dense(77, input_shape=input_shape, activation="relu", kernel_initializer='he_uniform')(X_input)
    # X = Dense(137, activation="relu", kernel_initializer='he_uniform')(X)
    X = Dense(77, activation="relu", kernel_initializer='he_uniform')(X)
    # Output Layer with # of actions: 2 nodes (left, right)
    X = Dense(action_space, activation="linear", kernel_initializer='he_uniform')(X)

    # With LSTM layers
    # X = Dense(57, input_shape=input_shape, activation="relu", kernel_initializer='he_uniform')(X_input)
    # X = tf.expand_dims(X, axis=-1)
    # X = LSTM(57, return_sequences=True)(X)
    # X = LSTM(57)(X)
    # X = Dense(57, activation="relu", kernel_initializer='he_uniform')(X)
    # X = Dense(action_space, activation="linear", kernel_initializer='he_uniform')(X)

    # Hidden layer with 75 nodes
    # X = Dense(75, activation="relu", kernel_initializer='he_uniform')(X)

    # X = LSTM(75)(X)

    # Hidden layer with 64 nodes
    # X = Dense(64, activation="relu", kernel_initializer='he_uniform')(X)

    model = Model(inputs=X_input, outputs=X)
    model.compile(loss="mse", optimizer=Adam(lr=0.001), metrics=["accuracy"])

    model.summary()
    return model


def get_host_cpu_util(host):
    global cpu_util_dict
    try:
        host_cpu_util = requests.get(PROMETHEUS + '/api/v1/query', params={
            'query': 'avg without (cpu) (1-(rate(node_cpu_seconds_total{job="node-exporter", '
                     'instance="' + host + '", mode="idle"}[10s]) ))'})

        results = host_cpu_util.json()['data']['result']

        rate = 0
        for result in results:
            rate = float('{value[1]}'.format(**result))
            if rate is None or math.isnan(rate):
                rate = 0

        r = {NODE_MAP_IP_NAME[host]: rate * 100}
        cpu_util_dict.update(r)
        return rate * 100
    except requests.exceptions.RequestException as e:
        rate = 0
        r = {NODE_MAP_IP_NAME[host]: rate * 100}
        cpu_util_dict.update(r)
        return 0


def get_fn_latency_ratio():
    ratio = 0
    # checked_fns = {"float": 0, "matrix": 0, "load": 0, "dd": 0, "load2": 0, "dd2": 0}
    index = 0
    for fn in FN_ARRAY:

        try:
            fn_avg_response_time = requests.get(PROMETHEUS + '/api/v1/query', params={
                'query': 'sum(rate(function_duration_seconds_sum{function = "' + fn + '" }[3s]))/ sum(rate('
                                                                                      'function_duration_seconds_count{function = "' + fn + '" }[3s])) '})

            results = fn_avg_response_time.json()['data']['result']

            latency = 0
            for result in results:
                latency = float('{value[1]}'.format(**result))
                if latency is None or math.isnan(latency):
                    latency = 0
        except requests.exceptions.RequestException as e:
            latency = 0

        # for result in results:
        #     latency = float('{value[1]}'.format(**result))

        ratio += latency / FN_DEADLINE[index]
        index = index + 1

    return ratio


# Latency is calculated only from running pods(pending ones are ignored) - avg latency of apps which have pods in the selected node
# def get_fn_latency_ratio(node):
#     checked_fns = {"float": 0, "float2": 0, "load": 0, "load2": 0, "matrix": 0, "primary": 0}
#     ratio = 0
#     try:
#         running_pods = requests.get(PROMETHEUS + '/api/v1/query', params={
#             'query': 'kube_pod_info{namespace="default", node="' + node + '"}'})
#
#         pods = running_pods.json()['data']['result']
#
#         # print("%s: Running pods in node %s are %s" % (datetime.now(), str(node), str(pods)))
#
#         for pod in pods:
#             fn_name = (pod['metric']['pod']).split('-')[0]
#             if fn_name == "nginx" or checked_fns[fn_name] == 1:
#                 continue
#
#             try:
#                 running_status = requests.get(PROMETHEUS + '/api/v1/query', params={
#                     'query': 'kube_pod_status_phase{namespace="default", phase="Running", pod="' + (
#                         pod['metric']['pod']) + '"}'})
#
#                 status = running_status.json()['data']['result']
#                 pod_status = 0
#                 for stat in status:
#                     pod_status = float('{value[1]}'.format(**stat))
#                     if pod_status is None or math.isnan(pod_status):
#                         pod_status = 0
#             except requests.exceptions.RequestException as e:
#                 pod_status = 0
#
#             if pod_status == 0:
#                 continue
#
#             try:
#                 fn_avg_response_time = requests.get(PROMETHEUS + '/api/v1/query', params={
#                     'query': 'sum(rate(function_duration_seconds_sum{function = "' + fn_name + '" }[5s]))/ sum(rate('
#                                                                                                'function_duration_seconds_count{function = "' + fn_name + '" }[5s])) '})
#
#                 results = fn_avg_response_time.json()['data']['result']
#
#                 latency = 0
#                 for result in results:
#                     latency = float('{value[1]}'.format(**result))
#                     if latency is None or math.isnan(latency):
#                         latency = 0
#             except requests.exceptions.RequestException as e:
#                 latency = 0
#
#             checked_fns[fn_name] = 1
#             # print("%s: latency of pod %s is %s" % (datetime.now(), str((pod['metric']['pod'])), str(latency)))
#
#             ratio += latency / FN_DEADLINE[FN_NAME_INDEX[fn_name]]
#
#
#
#     except requests.exceptions.RequestException as e:
#         pass
#     # print("%s: Fn latency ratio for node %s is %s" % (datetime.now(), node, str(ratio)))
#     return ratio


def get_app_latency(fn):
    try:
        fn_avg_response_time = requests.get(PROMETHEUS + '/api/v1/query', params={
            'query': 'sum(rate(function_duration_seconds_sum{function = "' + fn + '" }[3s]))/ sum(rate('
                                                                                  'function_duration_seconds_count{function = "' + fn + '" }[3s])) '})

        results = fn_avg_response_time.json()['data']['result']

        latency = 0
        for result in results:
            latency = float('{value[1]}'.format(**result))
            if latency is None or math.isnan(latency):
                latency = 0
    except requests.exceptions.RequestException as e:
        latency = 0

    return latency / FN_DEADLINE_DICT[fn]


def get_fn_failure_rate(fn_array):
    ratio = 0
    index = 0
    for fn in fn_array:

        try:
            fn_failures = requests.get(PROMETHEUS + '/api/v1/query', params={
                'query': 'sum( increase(function_failures_total{function = "' + fn + '" }[10s]))'})

            failures = fn_failures.json()['data']['result']

            failure_rate = 0
            for failure in failures:
                failure_rate = float('{value[1]}'.format(**failure))
                if failure_rate is None or math.isnan(failure_rate):
                    failure_rate = 0

        except requests.exceptions.RequestException as e:
            failure_rate = 0

        # for failure in failures:
        #     failure_rate = float('{value[1]}'.format(**failures))

        # make sure the failure number and arrival number are done for the same duration (10s as of now)
        fn_arr_rate = get_fn_arrival_rate(fn)
        if fn_arr_rate > 0:
            ratio += (failure_rate / 10) / fn_arr_rate

        index = index + 1

    return ratio

    # try 30s


# def get_fn_arrival_rate(app):
#     try:
#         fn_arrival_rate = requests.get(PROMETHEUS + '/api/v1/query', params={
#             'query': 'sum(increase(function_calls_total{function = "' + app + '" }[10s]))'})
#
#         results = fn_arrival_rate.json()['data']['result']
#
#         rate = 0
#         for result in results:
#             rate = float('{value[1]}'.format(**result))
#             if rate is None or math.isnan(rate):
#                 rate = 0
#
#     except requests.exceptions.RequestException as e:
#         rate = 0
#
#     return rate / 10

def get_fn_arrival_rate():
    fn_arrival_rate_dict = {}

    for fn in FN_ARRAY:
        try:
            fn_arrival_rate = requests.get(PROMETHEUS + '/api/v1/query', params={
                'query': 'sum(increase(function_calls_total{function = "' + fn + '" }[10s]))'})

            results = fn_arrival_rate.json()['data']['result']

            rate = 0
            for result in results:
                rate = float('{value[1]}'.format(**result))
                if rate is None or math.isnan(rate):
                    rate = 0

        except requests.exceptions.RequestException as e:
            rate = 0

        # print("fn %s has an arrival rate of %s" % (fn, str(rate)))
        r = {fn: rate / 10}
        fn_arrival_rate_dict.update(r)

    return fn_arrival_rate_dict


def get_active_nodes(pod_sch):
    active_nodes = 0
    global active_node_list
    active_node_list = []
    active_fns = {"float": 0, "float2": 0, "load": 0, "load2": 0, "matrix": 0, "primary": 0}
    checked_fns = {"float": False, "float2": False, "load": False, "load2": False, "matrix": False, "primary": False}
    for node in NODE_NAME:
        # cpu_util = get_host_cpu_util(NODE_MAP_NAME_IP[node])
        # if cpu_util > 5:
        #     active_nodes += 1
        # kube pod info picks pods in  pending state too - therefore nodes with pending pods of fns with incoming requests are considered active too
        try:
            running_pods = requests.get(PROMETHEUS + '/api/v1/query', params={
                'query': 'kube_pod_info{namespace="default", node="' + node + '"}'})

            pods = running_pods.json()['data']['result']

            # print("%s: Pods in node %s are %s" % (datetime.now(), node, str(pods)))

            for pod in pods:
                fn_name = (pod['metric']['pod']).split('-')[0]

                if fn_name == "nginx" or fn_name == "test" or fn_name == "my" or (pod['metric']['pod']) == pod_sch:
                    continue

                if active_fns[fn_name] == 0 and checked_fns[fn_name] is False:
                    # Make sure to have a suitable duration in this query

                    try:
                        fn_arrival_rate = requests.get(PROMETHEUS + '/api/v1/query', params={
                            'query': 'sum(increase(function_calls_total{function = "' + fn_name + '" }[10s]))'})

                        results = fn_arrival_rate.json()['data']['result']

                        fn_requests = 0
                        for result in results:
                            fn_requests = float('{value[1]}'.format(**result))
                            if fn_requests is None or math.isnan(fn_requests):
                                fn_requests = 0

                    except requests.exceptions.RequestException as e:
                        fn_requests = 0

                    # print("%s: Fn requests for fn %s are %s" % (datetime.now(), fn_name, str(fn_requests)))

                    # for result in results:
                    #     fn_requests = float('{value[1]}'.format(**result))
                    checked_fns[fn_name] = True
                    if fn_requests > 0:
                        active_nodes += 1
                        active_node_list.append(node)
                        # print("Active node>>> ", node)
                        active_fns[fn_name] = 1
                        break
                else:
                    if active_fns[fn_name] > 0:
                        active_nodes += 1
                        active_node_list.append(node)
                        # print("Node %s is Active since fn %s is in active list " % (str(node), fn_name))
                        break

        except requests.exceptions.RequestException as e:
            pass

    return active_node_list


def get_active_node_no():
    a_nodes = 0
    # global active_node_list
    active_n_list = []
    a_fns = {"float": 0, "load": 0, "matrix": 0, "primary": 0, "dd": 0, "imageprocess": 0, "imageresize": 0,
             "todoadditem": 0, "tododelitem": 0, "todogitem": 0, "todolistitem": 0, "todoupitem": 0, "uploadimage": 0,
             "uploadvideo": 0, "video": 0}
    c_fns = {"float": False, "load": False, "matrix": False, "primary": False, "dd": False, "imageprocess": False,
             "imageresize": False, "todoadditem": False, "tododelitem": False, "todogitem": False,
             "todolistitem": False, "todoupitem": False, "uploadimage": False, "uploadvideo": False, "video": False}
    for node in NODE_NAME:
        # cpu_util = get_host_cpu_util(NODE_MAP_NAME_IP[node])
        # if cpu_util > 5:
        #     active_nodes += 1
        # kube pod info picks pods in  pending state too - therefore nodes with pending pods of fns with incoming requests are considered active too
        try:
            running_pods = requests.get(PROMETHEUS + '/api/v1/query', params={
                'query': 'kube_pod_info{namespace="default", node="' + node + '"}'})

            pods = running_pods.json()['data']['result']

            # print("%s: Pods in node %s are %s" % (datetime.now(), node, str(pods)))

            for pod in pods:
                fn_name = (pod['metric']['pod']).split('-')[0]

                if fn_name == "nginx" or fn_name == "test" or fn_name == "my":
                    continue

                if a_fns[fn_name] == 0 and c_fns[fn_name] is False:
                    # Make sure to have a suitable duration in this query

                    try:
                        fn_arrival_rate = requests.get(PROMETHEUS + '/api/v1/query', params={
                            'query': 'sum(increase(function_calls_total{function = "' + fn_name + '" }[10s]))'})

                        results = fn_arrival_rate.json()['data']['result']

                        fn_requests = 0
                        for result in results:
                            fn_requests = float('{value[1]}'.format(**result))
                            if fn_requests is None or math.isnan(fn_requests):
                                fn_requests = 0

                    except requests.exceptions.RequestException as e:
                        fn_requests = 0

                    # print("%s: Fn requests for fn %s are %s" % (datetime.now(), fn_name, str(fn_requests)))

                    # for result in results:
                    #     fn_requests = float('{value[1]}'.format(**result))
                    c_fns[fn_name] = True
                    if fn_requests > 0:
                        a_nodes += 1
                        active_n_list.append(node)
                        # active_node_list.append(node)
                        # print("Active node>>> ", node)
                        a_fns[fn_name] = 1
                        break
                else:
                    if a_fns[fn_name] > 0:
                        a_nodes += 1
                        active_n_list.append(node)
                        # active_node_list.append(node)
                        # print("Node %s is Active since fn %s is in active list " % (str(node), fn_name))
                        break

        except requests.exceptions.RequestException as e:
            pass

    return a_nodes, active_n_list


def get_act_no_no():
    global node_count
    global act_nodes
    global act_node_cost
    node_count = 0
    # average_act_nodes = 0
    act_nodes = 0
    act_node_cost = 0
    while True:
        # count += 1
        a_n_no, a_node_list = get_active_node_no()
        if a_n_no > 0:
            act_nodes += a_n_no
            node_count += 1
            for node in a_node_list:
                act_node_cost += NODE_HR_RATE[node] / 3600 * 2

            # print("Active nodes are: %s, thus node cost is %s" % (str(a_node_list), str(act_node_cost)))
        # act_nodes += get_active_node_no()
        # average_act_nodes = act_nodes / count
        global stop_an_thread
        if stop_an_thread:
            break
        time.sleep(2)
    return act_nodes / node_count, node_count, act_node_cost


def get_node_pod_no(n):
    try:
        running_pods = requests.get(PROMETHEUS + '/api/v1/query', params={
            'query': 'kube_pod_info{namespace="default", node="' + n + '"}'})

        pods = running_pods.json()['data']['result']

        count = len(pods)
        print("no of pods is %s" % str(count))
    except requests.exceptions.RequestException as e:
        pass

    return count


def get_node_status(n, fn):
    # active_nodes = 0\
    global active_nodes
    global active_node_list
    node_active = 0
    node = NODE_MAP_IP_NAME[n]
    fn_replica_count_in_node = 0
    # for node in NODE_NAME:
    # cpu_util = get_host_cpu_util(NODE_MAP_NAME_IP[node])
    # if cpu_util > 5:
    #     active_nodes += 1
    # kube pod info picks pods in  pending state too - therefore nodes with pending pods of fns with incoming requests are considered active too
    try:
        running_pods = requests.get(PROMETHEUS + '/api/v1/query', params={
            'query': 'kube_pod_info{namespace="default", node="' + node + '"}'})

        pods = running_pods.json()['data']['result']

        # print("%s: Pods in node %s are %s" % (datetime.now(), node, str(pods)))

        for pod in pods:
            fn_name = (pod['metric']['pod']).split('-')[0]

            if fn_name == "nginx" or fn_name == "test" or fn_name == "my":
                continue
            if fn_name == fn:
                fn_replica_count_in_node += 1

            if active_fns[fn_name] == 0 and checked_fns[fn_name] is False:
                # Make sure to have a suitable duration in this query

                try:
                    fn_arrival_rate = requests.get(PROMETHEUS + '/api/v1/query', params={
                        'query': 'sum(increase(function_calls_total{function = "' + fn_name + '" }[10s]))'})

                    results = fn_arrival_rate.json()['data']['result']

                    fn_requests = 0
                    for result in results:
                        fn_requests = float('{value[1]}'.format(**result))
                        if fn_requests is None or math.isnan(fn_requests):
                            fn_requests = 0

                except requests.exceptions.RequestException as e:
                    fn_requests = 0

                # print("%s: Fn requests for fn %s are %s" % (datetime.now(), fn_name, str(fn_requests)))

                # for result in results:
                #     fn_requests = float('{value[1]}'.format(**result))
                checked_fns[fn_name] = True
                if fn_requests > 0:
                    active_nodes += 1
                    active_node_list.append(node)
                    node_active = 1
                    # print("Active node>>> ", node)
                    active_fns[fn_name] = 1
                    break
            else:
                if active_fns[fn_name] > 0:
                    active_nodes += 1
                    active_node_list.append(node)
                    node_active = 1
                    # print("Node %s is Active since fn %s is in active list " % (str(node), fn_name))
                    break

    except requests.exceptions.RequestException as e:
        pass

    return node_active, fn_replica_count_in_node


# fill replay buffer with n-step  buffer data
def get_n_step_info(n_s_buffer, gam):
    """Return n step reward, next state, and done."""
    # info of the last transition
    reward, next_state, done = n_s_buffer[-1][-3:]

    for transition in reversed(list(n_s_buffer)[:-1]):
        r, n_s, d = transition[-3:]

        reward = r + gam * reward * (1 - d)
        next_state, done = (n_s, d) if d else (next_state, done)

    return reward, next_state, done


# def get_reward(t_name, action, step, pod_name, ep, sch_data, wb, active_node_p, total_s, thread_no):
#     # while counter:
#     #     if exitFlag:
#     #         threadName.exit()
#     #     time.sleep(delay)
#     #     print("%s: %s" % (threadName, time.ctime(time.time())))
#     #     counter -= 1
#
#     time.sleep(4)
#     initial_lat = get_fn_latency_ratio(NODE_MAP_IP_NAME[action])
#
#     ###################################
#     # try:
#     #     pod_running = requests.get(PROMETHEUS + '/api/v1/query', params={
#     #         'query': 'kube_pod_status_phase{phase="Running",  namespace="default",pod="' + pod_name + '"}'})
#     #
#     #     # print(response.json())
#     #     results = pod_running.json()['data']['result']
#     #     fn_requests = 0
#     #     for result in results:
#     #         fn_requests = float('{value[1]}'.format(**result))
#     #         if fn_requests is None or math.isnan(fn_requests):
#     #             fn_requests = 0
#     #
#     #     print("At initial fn latency running status of pod :" + pod_name + " " + str(fn_requests))
#     #
#     # except requests.exceptions.RequestException as e:
#     #     fn_requests = 0
#
#     ##################################################
#
#     time.sleep(1)
#     print("Thread " + t_name + " *****************Reward function")
#     y = 0
#     pod_start = False
#     pod_terminated = False
#     while not pod_start:
#         if y != 0:
#             time.sleep(1)
#         time_now = datetime.now()
#         print(str(y) + " Time: ", time_now)
#
#         try:
#             pod_running = requests.get(PROMETHEUS + '/api/v1/query', params={
#                 'query': 'kube_pod_status_phase{phase="Running",  namespace="default",pod="' + pod_name + '"}'})
#
#             # print(response.json())
#             results = pod_running.json()['data']['result']
#
#             if not pod_running or pod_running is None:
#                 pod_terminated = True
#                 break
#
#             # print(results)
#             for result in results:
#                 fn_requests = float('{value[1]}'.format(**result))
#                 if fn_requests is None or math.isnan(fn_requests):
#                     # print("Pod: " + pod_name + " is terminated")
#                     pod_terminated = True
#                     break
#                 elif fn_requests == 1:
#                     # print("result: ", '{value[1]}'.format(**result))
#                     # print("Pod: " + pod_name + " is running")
#                     pod_start = True
#                 else:
#                     print("result: ", '{value[1]}'.format(**result))
#                     # print("Pod: " + pod_name + " is NOT running")
#
#         except requests.exceptions.RequestException as e:
#             pass
#
#         if pod_terminated:
#             break
#         else:
#             y = y + 1
#             # print("pod_running :" + pod_name + " " + str(pod_start))
#             if y > 40:
#                 pod_terminated = True
#                 break
#
#     # print("Adding thread " + str(t_name) + " finished queue")
#     # thread_finish_q.put(t_name)
#
#     if pod_terminated:
#         reward = -50
#         time_now = datetime.now()
#         sch_data.write(step, 10, time_now)
#         # print("Time: ", time_now)
#         print("Pod: " + pod_name + " is terminated so reward: ", reward)
#
#         pending_pod_list.remove(pod_name)
#
#
#
#     else:
#         pending_pod_list.remove(pod_name)
#         time_now = datetime.now()
#         sch_data.write(step, 10, time_now)
#         # print("Time: ", time_now)
#         # print("Thread " + t_name + " Starting sleep3")
#         time.sleep(2)
#         time_now = datetime.now()
#         # print("Time: ", time_now)
#         print("Thread " + t_name + " *****************Starting Reward calculation")
#         reward_fn_latency = get_fn_latency_ratio(NODE_MAP_IP_NAME[action]) - initial_lat
#
#         global function_latency
#         function_latency += reward_fn_latency
#
#
#         # if reward_fn_latency < 0:
#         #     reward_fn_latency = 0
#
#         # print("fn_latency_ratio: ", reward_fn_latency)
#         reward_fn_failure_rate = get_fn_failure_rate(FN_ARRAY)
#         # print("fn_failure_rate: ", reward_fn_failure_rate)
#         cpu_util_of_selected_node = get_host_cpu_util(action)
#         # print("CPU_util_of_selected_node: ", cpu_util_of_selected_node)
#
#         if cpu_util_of_selected_node > 80:
#             cpu_penalty = cpu_util_of_selected_node / 100
#
#         else:
#             cpu_penalty = 0
#
#         # print("CPU penalty: ", cpu_penalty)
#         # active_node_no = get_active_nodes()
#         # print("active_node_no: ", active_node_no)
#
#         # reward = - (reward_fn_latency + reward_fn_failure_rate + cpu_penalty + active_node_p)
#         reward = - reward_fn_latency * 5
#
#         sch_data.write(step, 11, reward_fn_latency)
#         sch_data.write(step, 12, reward_fn_failure_rate)
#         sch_data.write(step, 13, cpu_penalty)
#         sch_data.write(step, 14, active_node_p)
#
#         # print("reward: ", reward)
#
#     global episodic_reward
#     episodic_reward += reward
#
#
#
#     global act_node_penalty
#     act_node_penalty += active_node_p
#
#     # appending the reward correctly
#
#     # if reward_q.qsize() == 0
#     # print("REWARD queue size: " + str(reward_q.qsize()))
#
#     # print("ThreadIDq: ", threadID_q)
#     # print("Rewardq: ", reward_q)
#     time_now = datetime.now()
#
#     if threadID_q.qsize() != 0:
#         if threadID_q.queue[0] == thread_no:
#             # print("%s : since thread no is %s and threadID_q's first value is %s adding the reward to queue " % (str(time_now), str(thread_no), threadID_q.queue[0]))
#             reward_q.put(reward)
#
#         else:
#             # print("since thread no is %s and threadID_q's first value is %s adding the reward to DICT " % (str(thread_no), threadID_q.queue[0]))
#             d = {thread_no: reward}
#             reward_dict.update(d)
#             # print("Reward dict: " + str(reward_dict))
#
#     try:
#         # print("Writing reward to excel")
#         sch_data.write(step, 6, reward)
#         # print("Wrote reward to excel")
#         wb.save("Episodic_Data" + str(ep) + ".csv")
#         # print("Saved reward to excel")
#     except Exception as inst:
#         # print(type(inst))  # the exception instance
#         # print(inst.args)  # arguments stored in .args
#         print(inst)
#
#     # print("After thread: " + t_name)
#     # print("state: ")
#     # for elem in list(state_q.queue):
#     #     print(elem)
#     # print("reward: ")
#     # for elem in list(reward_q.queue):
#     #     print(elem)
#     # print("action: ")
#     # for elem in list(action_q.queue):
#     #     print(elem)
#
#     while threadID_q.qsize() != 0 and state_q.qsize() != 0 and reward_q.qsize() != 0 and action_q.qsize() != 0 and next_state_q.qsize() != 0 and done_q.qsize() != 0:
#
#         print("Appending to buffer")
#         memory.append(
#             (state_q.get(), action_q.get(), reward_q.get(), next_state_q.get(), done_q.get()))
#         # print("n_step_buffer length: ", len(n_step_buffer))
#         threadID_q.get()
#
#         # print("n_step_buffer: %s " % (str(n_step_buffer)))
#
#         # if len(n_step_buffer) == n_step:  # fill the n-step buffer for the first translation
#         #     # add a multi step transition
#         #     reward, next_obs, done = get_n_step_info(n_step_buffer, gamma_nstep)
#         #     obs, action = n_step_buffer[0][:2]
#         #
#         #     print("Appending to memory")
#         #     memory.append((obs, action, reward, next_obs, done))
#
#         print("memory size : ", len(memory))
#         # print("memory : %s " % (str(memory)))
#
#         if threadID_q.qsize() != 0:
#             for key in reward_dict:
#                 if key == threadID_q.queue[0]:
#                     # print("Adding reward of thread %s to queue after memory update" % str(key))
#                     reward_q.put(reward_dict[key])
#                     reward_dict.pop(key)
#                     # print("Now thread dict %s " % str(reward_dict))
#                     break
#         # print("Memory now: ", memory)
#
#     global epsilon
#     if len(memory) > train_start:
#         if total_s % 5 == 0:
#             if epsilon > epsilon_min:
#                 epsilon *= epsilon_decay
#                 print("changed Epsilon: " + str(epsilon))


##**
def set_reward_new(t_name, action, step, pod_name, ep, sch_data, wb, active_node_p, total_s, thread_no, u_p):
    # while counter:
    #     if exitFlag:
    #         threadName.exit()
    #     time.sleep(delay)
    #     print("%s: %s" % (threadName, time.ctime(time.time())))
    #     counter -= 1

    # app = pod_name.split('-')[0]
    time.sleep(3)
    # tail_lat = get_app_latency(app)
    # tail_lat = round(get_fn_latency_ratio(NODE_MAP_IP_NAME[action]), 2)
    # tail_lat = get_fn_latency_ratio()
    global vm_up_time_prev
    global vm_up_time_cost_prev
    global initial_lat
    global prev_action
    global host_p

    # global prev_app
    if step > 1:
        # tail_lat_prev_action = round(get_fn_latency_ratio(NODE_MAP_IP_NAME[prev_action]), 2)
        # cpu_util_prev_action = round(get_host_cpu_util(prev_action)/10, 1)
        # tail_lat_prev_app = get_app_latency(prev_app)
        tail_lat_prev_app = round(get_fn_latency_ratio(), 2)
        # vm_up_time = act_nodes * 2

    prev_action = action
    vm_up_time = act_nodes * 2
    vm_up_time_cost = act_node_cost
    time.sleep(2)

    # prev_app = app
    # tail_lat = get_fn_latency_ratio()

    ###################################
    # try:
    #     pod_running = requests.get(PROMETHEUS + '/api/v1/query', params={
    #         'query': 'kube_pod_status_phase{phase="Running",  namespace="default",pod="' + pod_name + '"}'})
    #
    #     # print(response.json())
    #     results = pod_running.json()['data']['result']
    #     fn_requests = 0
    #     for result in results:
    #         fn_requests = float('{value[1]}'.format(**result))
    #         if fn_requests is None or math.isnan(fn_requests):
    #             fn_requests = 0
    #
    #     print("At initial fn latency running status of pod :" + pod_name + " " + str(fn_requests))
    #
    # except requests.exceptions.RequestException as e:
    #     fn_requests = 0

    ##################################################

    print("Step " + t_name + " *****************Reward function")
    y = 0
    pod_start = False
    pod_terminated = False
    while not pod_start:
        if y != 0:
            time.sleep(1)
        time_now = datetime.now()
        print(str(y) + " Time: ", time_now)

        try:
            pod_running = requests.get(PROMETHEUS + '/api/v1/query', params={
                'query': 'kube_pod_status_phase{phase="Running",  namespace="default",pod="' + pod_name + '"}'})

            # print(response.json())
            results = pod_running.json()['data']['result']

            if not pod_running or pod_running is None:
                pod_terminated = True
                break

            # print(results)
            for result in results:
                fn_requests = float('{value[1]}'.format(**result))
                if fn_requests is None or math.isnan(fn_requests):
                    # print("Pod: " + pod_name + " is terminated")
                    pod_terminated = True
                    break
                elif fn_requests == 1:
                    # print("result: ", '{value[1]}'.format(**result))
                    print("Pod: " + pod_name + " is running")
                    pod_start = True
                else:
                    # print("result: ", '{value[1]}'.format(**result))
                    print("Pod: " + pod_name + " is NOT running")

        except requests.exceptions.RequestException as e:
            pass

        if pod_terminated:
            break
        else:
            y = y + 1
            print("pod_running :" + pod_name + " " + str(pod_start))
            if y > 15:
                pod_terminated = True
                break

    # print("Adding thread " + str(t_name) + " finished queue")
    # thread_finish_q.put(t_name)

    try:
        time_now = datetime.now()
        sch_data.write(step, 10, time_now)
    except Exception as inst:
        print(inst)

    if pod_terminated:
        # reward = -5
        # time_now = datetime.now()
        # sch_data.write(step, 10, time_now)
        # print("Time: ", time_now)
        print("Pod: " + pod_name + " is terminated so reward: -5 ")

        # pending_pod_list.remove(pod_name)

    # else:

    # pending_pod_list.remove(pod_name)
    # time_now = datetime.now()
    # sch_data.write(step, 10, time_now)
    # print("Time: ", time_now)
    # print("Thread " + t_name + " Starting sleep3")

    # time.sleep(2)
    # time_now = datetime.now()

    # print("Time: ", time_now)

    # reward_fn_latency = get_fn_latency_ratio(NODE_MAP_IP_NAME[action]) - initial_lat

    if step > 1:
        if initial_lat == -5:
            # reward_fn_latency = -5
            reward_fn_latency = 0
            reward_vm_up_time = 0
            reward_vm_up_time_cost = 0
            # cpu_penalty = 0

        else:
            # reward_fn_latency = tail_lat_prev_action - initial_lat
            reward_fn_latency = tail_lat_prev_app
            reward_vm_up_time = vm_up_time - vm_up_time_prev
            reward_vm_up_time_cost = vm_up_time_cost - vm_up_time_cost_prev
            # reward_fn_latency = tail_lat_prev_action - initial_lat
            # if cpu_util_prev_action >= 7:
            #     cpu_penalty = cpu_util_prev_action
            #
            # else:
            #     cpu_penalty = 0

            # reward_fn_latency = host_p
            if reward_fn_latency < 0:
                reward_fn_latency = 0

    # host_p = u_p
    if pod_terminated:
        initial_lat = -5
    else:
        # initial_lat = tail_lat
        initial_lat = 0
        vm_up_time_prev = vm_up_time
        vm_up_time_cost_prev = vm_up_time_cost
    if step > 1:
        print("Step " + t_name + " *****************Starting Reward calculation for action before")

        if abs(reward_fn_latency) > 40:
            if reward_fn_latency < 0:
                reward_fn_latency = -40
            else:
                reward_fn_latency = 40

        global function_latency
        function_latency += reward_fn_latency

        global total_vm_time_diff
        total_vm_time_diff += reward_vm_up_time

        global total_vm_cost_diff
        total_vm_cost_diff += reward_vm_up_time_cost

        global cpu_util_penalty
        # cpu_util_penalty += cpu_penalty

        # reward = - (reward_fn_latency + cpu_penalty) * 5

        # reward = - reward_fn_latency
        x = 0
        y = 1 - x
        # reward = - (reward_vm_up_time / 200 * x + reward_fn_latency / 30 * y)
        # vm cost multiplying by 100 to make that value as significant as the latency value
        reward = - (reward_vm_up_time_cost * 100 * x + reward_fn_latency / 10 * y)

        sch_data.write(step - 1, 11, reward_fn_latency)
        sch_data.write(step - 1, 15, reward_vm_up_time_cost)
        sch_data.write(step - 1, 16, reward_vm_up_time)

        # sch_data.write(step - 1, 13, cpu_penalty)

        reward_q.put(reward)
        global episodic_reward
        episodic_reward += reward

        # if reward_fn_latency < 0:
        #     reward_fn_latency = 0

        # print("fn_latency_ratio: ", reward_fn_latency)
        # reward_fn_failure_rate = get_fn_failure_rate(FN_ARRAY)
        # # print("fn_failure_rate: ", reward_fn_failure_rate)
        # cpu_util_of_selected_node = get_host_cpu_util(action)
        # # print("CPU_util_of_selected_node: ", cpu_util_of_selected_node)
        #
        # if cpu_util_of_selected_node > 80:
        #     cpu_penalty = cpu_util_of_selected_node / 100
        #
        # else:
        #     cpu_penalty = 0

        # print("CPU penalty: ", cpu_penalty)
        # active_node_no = get_active_nodes()
        # print("active_node_no: ", active_node_no)

        # reward = - (reward_fn_latency + reward_fn_failure_rate + cpu_penalty + active_node_p)

        # sch_data.write(step, 12, reward_fn_failure_rate)
        # sch_data.write(step, 13, cpu_penalty)

        # print("reward: ", reward)

    global act_node_penalty
    act_node_penalty += active_node_p

    # appending the reward correctly

    # if reward_q.qsize() == 0
    # print("REWARD queue size: " + str(reward_q.qsize()))

    # print("ThreadIDq: ", threadID_q)
    # print("Rewardq: ", reward_q)
    # time_now = datetime.now()

    # if threadID_q.qsize() != 0:
    #     if threadID_q.queue[0] == thread_no:
    #         # print("%s : since thread no is %s and threadID_q's first value is %s adding the reward to queue " % (str(time_now), str(thread_no), threadID_q.queue[0]))
    #         reward_q.put(reward)
    #
    #     else:
    #         # print("since thread no is %s and threadID_q's first value is %s adding the reward to DICT " % (str(thread_no), threadID_q.queue[0]))
    #         d = {thread_no: reward}
    #         reward_dict.update(d)
    # print("Reward dict: " + str(reward_dict))

    try:
        # print("Writing reward to excel")
        if step > 1:
            sch_data.write(step - 1, 6, reward)
        sch_data.write(step, 14, active_node_p)
        # print("Wrote reward to excel")
        wb.save("episodic_data/Episodic_Data" + str(ep) + ".csv")
        # print("Saved reward to excel")
    except Exception as inst:
        # print(type(inst))  # the exception instance
        # print(inst.args)  # arguments stored in .args
        print(inst)

    # print("After thread: " + t_name)
    # print("state: ")
    # for elem in list(state_q.queue):
    #     print(elem)
    # print("reward: ")
    # for elem in list(reward_q.queue):
    #     print(elem)
    # print("action: ")
    # for elem in list(action_q.queue):
    #     print(elem)

    while state_q.qsize() != 0 and reward_q.qsize() != 0 and action_q.qsize() != 0 and next_state_q.qsize() != 0 and done_q.qsize() != 0:
        print("Appending to buffer")
        # memory.append(
        #     (state_q.get(), action_q.get(), reward_q.get(), next_state_q.get(), done_q.get()))

        n_step_buffer.append(
            (state_q.get(), action_q.get(), reward_q.get(), next_state_q.get(), done_q.get()))
        print("n-step buffer length: ", len(n_step_buffer))

        if len(n_step_buffer) == n_step:  # fill the n-step buffer for the first translation
            # add a multi step transition
            reward, next_obs, done = get_n_step_info(n_step_buffer, gamma_nstep)
            obs, action = n_step_buffer[0][:2]

            memory.append((obs, action, reward, next_obs, done))
        # print("n_step_buffer length: ", len(n_step_buffer))
        # threadID_q.get()

        # print("n_step_buffer: %s " % (str(n_step_buffer)))

        # if len(n_step_buffer) == n_step:  # fill the n-step buffer for the first translation
        #     # add a multi step transition
        #     reward, next_obs, done = get_n_step_info(n_step_buffer, gamma_nstep)
        #     obs, action = n_step_buffer[0][:2]
        #
        #     print("Appending to memory")
        #     memory.append((obs, action, reward, next_obs, done))

        print("memory size : ", len(memory))
        # print("memory : %s " % (str(memory)))

        # if threadID_q.qsize() != 0:
        #     for key in reward_dict:
        #         if key == threadID_q.queue[0]:
        #             # print("Adding reward of thread %s to queue after memory update" % str(key))
        #             reward_q.put(reward_dict[key])
        #             reward_dict.pop(key)
        #             # print("Now thread dict %s " % str(reward_dict))
        #             break
        # print("Memory now: ", memory)

    global epsilon
    if len(memory) > train_start:
        if total_s % 2 == 0:
            if epsilon > epsilon_min:
                epsilon *= epsilon_decay
                print("changed Epsilon: " + str(epsilon))

    time.sleep(2)


##**


def set_reward_test(t_name, action, step, pod_name, ep, sch_data, wb, active_node_p, total_s, thread_no, u_p):
    time.sleep(3)
    # tail_lat = get_app_latency(app)
    # tail_lat = round(get_fn_latency_ratio(NODE_MAP_IP_NAME[action]), 2)
    # tail_lat = get_fn_latency_ratio()
    global initial_lat
    global prev_action
    global host_p
    global vm_up_time_prev
    global vm_up_time_cost_prev
    global act_node_cost
    # global prev_app
    if step > 1:
        # tail_lat_prev_action = round(get_fn_latency_ratio(NODE_MAP_IP_NAME[prev_action]), 2)
        # cpu_util_prev_action = round(get_host_cpu_util(prev_action)/10, 1)
        # tail_lat_prev_app = get_app_latency(prev_app)
        tail_lat_prev_app = round(get_fn_latency_ratio(), 2)

    prev_action = action
    vm_up_time = act_nodes * 2
    vm_up_time_cost = act_node_cost

    time.sleep(2)

    print("Step " + t_name + " *****************Reward function")
    y = 0
    pod_start = False
    pod_terminated = False
    while not pod_start:
        if y != 0:
            time.sleep(1)
        time_now = datetime.now()
        print(str(y) + " Time: ", time_now)

        try:
            pod_running = requests.get(PROMETHEUS + '/api/v1/query', params={
                'query': 'kube_pod_status_phase{phase="Running",  namespace="default",pod="' + pod_name + '"}'})

            # print(response.json())
            results = pod_running.json()['data']['result']

            if not pod_running or pod_running is None:
                pod_terminated = True
                break

            # print(results)
            for result in results:
                fn_requests = float('{value[1]}'.format(**result))
                if fn_requests is None or math.isnan(fn_requests):
                    # print("Pod: " + pod_name + " is terminated")
                    pod_terminated = True
                    break
                elif fn_requests == 1:
                    # print("result: ", '{value[1]}'.format(**result))
                    print("Pod: " + pod_name + " is running")
                    pod_start = True
                else:
                    # print("result: ", '{value[1]}'.format(**result))
                    print("Pod: " + pod_name + " is NOT running")

        except requests.exceptions.RequestException as e:
            pass

        if pod_terminated:
            break
        else:
            y = y + 1
            print("pod_running :" + pod_name + " " + str(pod_start))
            if y > 15:
                pod_terminated = True
                break

    # print("Adding thread " + str(t_name) + " finished queue")
    # thread_finish_q.put(t_name)

    try:
        time_now = datetime.now()
        sch_data.write(step, 10, time_now)
    except Exception as inst:
        print(inst)

    if pod_terminated:
        print("Pod: " + pod_name + " is terminated so reward: -5 ")

    if step > 1:
        if initial_lat == -5:
            # reward_fn_latency = -5
            reward_fn_latency = 0
            reward_vm_up_time = 0
            reward_vm_up_time_cost = 0
            # cpu_penalty = 0

        else:
            # reward_fn_latency = tail_lat_prev_action - initial_lat
            reward_fn_latency = tail_lat_prev_app
            reward_vm_up_time = vm_up_time - vm_up_time_prev
            reward_vm_up_time_cost = vm_up_time_cost - vm_up_time_cost_prev
            # reward_fn_latency = host_p
            if reward_fn_latency < 0:
                reward_fn_latency = 0

    # host_p = u_p
    if pod_terminated:
        initial_lat = -5
    else:
        # initial_lat = tail_lat
        initial_lat = 0
        vm_up_time_prev = vm_up_time
        vm_up_time_cost_prev = vm_up_time_cost
    if step > 1:
        print("Step " + t_name + " *****************Starting Reward calculation for action before")

        if abs(reward_fn_latency) > 40:
            if reward_fn_latency < 0:
                reward_fn_latency = -40
            else:
                reward_fn_latency = 40

        global function_latency
        function_latency += reward_fn_latency

        global total_vm_time_diff
        total_vm_time_diff += reward_vm_up_time

        global total_vm_cost_diff
        total_vm_cost_diff += reward_vm_up_time_cost

        global cpu_util_penalty
        # cpu_util_penalty += cpu_penalty

        # reward = - (reward_fn_latency + cpu_penalty) * 5

        # reward = - reward_fn_latency
        # reward = - reward_vm_up_time
        x = 0
        y = 1 - x
        # reward = - (reward_vm_up_time / 200 * x + reward_fn_latency / 30 * y)
        reward = - (reward_vm_up_time_cost * x + reward_fn_latency / 30 * y)

        sch_data.write(step - 1, 11, reward_fn_latency)
        sch_data.write(step - 1, 15, reward_vm_up_time_cost)
        sch_data.write(step - 1, 16, reward_vm_up_time)
        # sch_data.write(step - 1, 13, cpu_penalty)

        # reward_q.put(reward)
        global episodic_reward
        episodic_reward += reward

    global act_node_penalty
    act_node_penalty += active_node_p

    try:
        # print("Writing reward to excel")
        if step > 1:
            sch_data.write(step - 1, 6, reward)
        sch_data.write(step, 14, active_node_p)
        # print("Wrote reward to excel")
        wb.save("results/Evaluation_Data" + str(ep) + ".csv")
        # print("Saved reward to excel")
    except Exception as inst:
        # print(type(inst))  # the exception instance
        # print(inst.args)  # arguments stored in .args
        print(inst)

    time.sleep(2)


class epThread(threading.Thread):
    def __init__(self, thread_name, page, workbk, row, ep):
        threading.Thread.__init__(self)
        self.thread_name = thread_name
        self.sc_data = page
        self.wbook = workbk
        self.place = row
        self.epi = ep

    def run(self):
        # print("Starting " + self.thread_name)
        avg_active_nodes, active_node_count, total_node_cost = get_act_no_no()
        agent.tensorboard.update_stats(Average_Node_No=avg_active_nodes)
        agent.tensorboard.update_stats(VM_Up_Time=avg_active_nodes * active_node_count * 2)
        agent.tensorboard.update_stats(Active_Node_Cost=total_node_cost)
        self.sc_data.write(self.place, 4, avg_active_nodes)
        self.wbook.save("episodic_data/Episodic_Data" + str(self.epi) + ".csv")
        print("Exiting " + self.thread_name)


class epThread_test(threading.Thread):
    def __init__(self, thread_name, page, workbk, row, ep):
        threading.Thread.__init__(self)
        self.thread_name = thread_name
        self.sc_data = page
        self.wbook = workbk
        self.place = row
        self.epi = ep

    def run(self):
        print("Starting " + self.thread_name)
        avg_active_nodes, active_node_count, total_node_cost = get_act_no_no()
        agent.tensorboard.update_stats(Average_Node_No=avg_active_nodes)
        agent.tensorboard.update_stats(VM_Up_Time=avg_active_nodes * active_node_count * 2)
        agent.tensorboard.update_stats(Active_Node_Cost=total_node_cost)
        self.sc_data.write(self.place, 4, avg_active_nodes)
        self.wbook.save("results/Evaluation_Data" + str(self.epi) + ".csv")
        print("Exiting " + self.thread_name)


# class myThread(threading.Thread):
#     def __init__(self, threadID, name, node, step, pod_n, epi, sheet, wrkbk, node_penalty, total_steps):
#         threading.Thread.__init__(self)
#         self.threadID = threadID
#         self.name = name
#         self.action = node
#         self.time_step = step
#         self.pod = pod_n
#         self.episode = epi
#         self.sch_data = sheet
#         self.wb = wrkbk
#         # self.initial_latency = node_latency
#         self.node_p = node_penalty
#         self.t_steps = total_steps
#         # self.done = done_status
#
#     def run(self):
#         print("Starting " + self.name)
#         get_reward(self.name, self.action, self.time_step, self.pod, self.episode, self.sch_data, self.wb, self.node_p,
#                    self.t_steps, self.threadID)
#         print("Exiting " + self.name)


# Own Tensorboard class
class ModifiedTensorBoard(TensorBoard):
    # Overriding init to set initial step and writer (we want one log file for all .fit() calls)
    def __init__(self, name, **kwargs):
        super().__init__(**kwargs)
        self.step = 1
        self.writer = tf.summary.create_file_writer(self.log_dir)
        self._log_write_dir = os.path.join(self.log_dir, name)

        # self._train_dir = 'train'

    # Overriding this method to stop creating default log writer
    def set_model(self, model):
        # self.model = model
        self.model = model
        self._log_write_dir = self.log_dir
        self._train_dir = os.path.join(self._log_write_dir, 'train')
        self._train_step = self.model._train_counter

        self._val_dir = os.path.join(self._log_write_dir, 'validation')
        self._val_step = self.model._test_counter
        # pass

    # Overrided, saves logs with our step number
    # (otherwise every .fit() will start writing from 0th step)
    def on_epoch_end(self, epoch, logs=None):
        self.update_stats(**logs)

    # Overrided
    # We train for one batch only, no need to save anything at epoch end
    def on_batch_end(self, batch, logs=None):
        pass

    # Overrided, so won't close writer
    def on_train_end(self, _):
        pass

    def on_train_batch_end(self, batch, logs=None):
        pass

    # Custom method for saving own metrics
    # Creates writer, writes custom metrics and closes writer
    def update_stats(self, **stats):
        self._write_logs(stats, self.step)

    def _write_logs(self, logs, index):
        with self.writer.as_default():
            for name, value in logs.items():
                tf.summary.scalar(name, value, step=index)
                self.step += 1
                self.writer.flush()


# tensorboard = TensorBoard(
#     log_dir=f"/home/ubuntu/work/logs/{MODEL_NAME}-{str(datetime.now())}",
#     histogram_freq=1,
#     write_images=True
# )
# keras_callbacks = [
#     tensorboard
# ]

# log_dir = f"/home/ubuntu/work/logs/{MODEL_NAME}-{int(time.time())}"
# # tensorboard_callback = tf.keras.callbacks.TensorBoard(log_dir=log_dir, histogram_freq=1)
# tensorboard_callback = tf.keras.callbacks.TensorBoard(log_dir=log_dir)


class DQNAgent:
    def __init__(self):
        # self.env = gym.make('CartPole-v1')
        # by default, CartPole-v1 has max episode steps = 500
        # Host CPU 0-100 percentage | Memory 0 100 percentage | BW (Transmit+Receive) 0 100 MB/s | Disk IO 0 100 MB/s
        # Host FN CPU 0-1 cores | Memory 0 100 MB | NW (tr+rec) 0 100 MB/s Arrival rate 0 30 requests/s | Replicas 0 6
        # FN present or not in the host 0 1
        # FN CPU (10 requests/s for 1s) Matrix: 0.05 | Float 0.15 | Load: 0.01 | DD: 0.09
        # FN Memory (10 requests/s for 1s) Matrix  38 MB | Float 38 MB | Load 38 MB | DD 126 MB
        # FN NW (10 requests/s for 1s)(receive+transmit)  | Load 50 (25+25) MB/s |
        # FN IO (10 requests/s for 1s)(read+write) DD 6 MB/s (0+6)
        # float matrix load dd
        # [Node 1 CPU memory, BW, Disk IO | Node 2 CPU, memory, BW, Disk IO,..........,|fn CPU, memory, BW, arrival rate, replica no| fn1 present, fn2 present.....|fn1 arrival rate, fn2 arrivalrate....| fn1 replica no, fn2 replica no....]
        self.observation_space = Box(low=np.array(
            [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0,
             0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0,
             0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]),
            high=np.array(
                [1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1,
                 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1,
                 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1,
                 1, 1, 1, 1, 1, 1, 1]),
            dtype=np.float32)
        self.state_size = self.observation_space.shape[0]
        self.action_space = Discrete(10)
        self.action_size = self.action_space.n
        print(self.state_size)
        print(self.action_size)
        print(self.observation_space.shape)

        self.EPISODES = 5000
        self.EPISODES_TEST = 100
        self.gamma = 0.95  # discount rate
        self.batch_size = 64
        self.update_rate = 100

        # print(self.env.observation_space.shape[0])
        # print(self.env.observation_space.shape)
        # print(self.env.observation_space)
        # state = self.env.reset()
        # print(state)

        # self.state_q = Queue()
        # self.reward_q = Queue()
        # self.action_q = Queue()
        # self.next_state_q = Queue()

        # self.epsilon = 0.8  # exploration rate
        # self.epsilon_min = 0.002
        # self.epsilon_decay = 0.998

        # self.train_start = 64

        # create main model
        self.main_network = OurModel(input_shape=(self.state_size,), action_space=self.action_size)
        self.target_network = OurModel(input_shape=(self.state_size,), action_space=self.action_size)
        self.target_network.set_weights(self.main_network.get_weights())

        # self.tensorboard = ModifiedTensorBoard(log_dir=f"logs/{MODEL_NAME}-{int(time.time())}")
        self.tensorboard = ModifiedTensorBoard(MODEL_NAME,
                                               log_dir="{}logs/{}-{}".format(PATH, MODEL_NAME, int(time.time())))

    def get_host_memory_util(self, host):

        try:
            host_memory_util = requests.get(PROMETHEUS + '/api/v1/query', params={
                'query': '((node_memory_MemTotal_bytes{instance="' + host + '",job="node-exporter"} - node_memory_MemFree_bytes{instance="' + host + '",job="node-exporter"} - node_memory_Cached_bytes{job="node-exporter",instance="' + host + '"} - node_memory_Buffers_bytes{instance="' + host + '",job="node-exporter"})/(node_memory_MemTotal_bytes{instance="' + host + '",job="node-exporter"}))'})

            results = host_memory_util.json()['data']['result']

            # for result in results:
            #     return float('{value[1]}'.format(**result))

            rate = 0
            for result in results:
                rate = float('{value[1]}'.format(**result))
                if rate is None or math.isnan(rate):
                    rate = 0

        except requests.exceptions.RequestException as e:
            rate = 0

        return rate * 100

    # Returned in MB/s
    def get_host_bw_util(self, host):

        try:
            host_nw_receive = requests.get(PROMETHEUS + '/api/v1/query', params={
                'query': 'sum(rate(node_network_receive_bytes_total{job="node-exporter",'
                         'instance="' + host + '"}[20s]))'})

            receive = host_nw_receive.json()['data']['result']

            receive_bw = 0
            for result in receive:
                receive_bw = float('{value[1]}'.format(**result))
                if receive_bw is None or math.isnan(receive_bw):
                    receive_bw = 0

        except requests.exceptions.RequestException as e:
            receive_bw = 0

        # for result in receive:
        #     receive_bw = float('{value[1]}'.format(**result))

        try:
            host_nw_transmit = requests.get(PROMETHEUS + '/api/v1/query', params={
                'query': 'sum(rate(node_network_transmit_bytes_total{job="node-exporter",'
                         'instance="' + host + '"}[20s]))'})

            transmit = host_nw_transmit.json()['data']['result']

            transmit_bw = 0
            for result in transmit:
                transmit_bw = float('{value[1]}'.format(**result))
                if transmit_bw is None or math.isnan(transmit_bw):
                    transmit_bw = 0

        except requests.exceptions.RequestException as e:
            transmit_bw = 0

        # for result in transmit:
        #     transmit_bw = float('{value[1]}'.format(**result))

        return (receive_bw + transmit_bw) / 1000000

    # Returned as MB/s
    def get_host_diskio_util(self, host):

        try:
            host_disk_read_rate = requests.get(PROMETHEUS + '/api/v1/query', params={
                'query': 'sum(rate(node_disk_read_bytes_total{job="node-exporter", instance="' + host + '", ''device=~"nvme.+|rbd.+|sd.+|vd.+|xvd.+|dm-.+|dasd.+"}[20s]))'})

            read = host_disk_read_rate.json()['data']['result']

            read_rate = 0
            for result in read:
                read_rate = float('{value[1]}'.format(**result))
                if read_rate is None or math.isnan(read_rate):
                    read_rate = 0

        except requests.exceptions.RequestException as e:
            read_rate = 0

        # for result in read:
        #     read_rate = float('{value[1]}'.format(**result))

        try:

            host_disk_write_rate = requests.get(PROMETHEUS + '/api/v1/query', params={
                'query': 'sum(rate(node_disk_written_bytes_total{job="node-exporter", instance="' + host + '", ''device=~"nvme.+|rbd.+|sd.+|vd.+|xvd.+|dm-.+|dasd.+"}[20s]))'})

            write = host_disk_write_rate.json()['data']['result']

            write_rate = 0
            for result in write:
                write_rate = float('{value[1]}'.format(**result))
                if write_rate is None or math.isnan(write_rate):
                    write_rate = 0

        except requests.exceptions.RequestException as e:
            write_rate = 0

        # for result in write:
        #     write_rate = float('{value[1]}'.format(**result))

        return (read_rate + write_rate) / 1000000

    # Returned as a ratio of a core
    def get_fn_cpu_request(self, app):

        try:
            fn_cpu_request = requests.get(PROMETHEUS + '/api/v1/query', params={
                'query': 'kube_pod_container_resource_requests{ namespace="default", container = "' + app + '", resource = '
                                                                                                            '"cpu"}'})

            results = fn_cpu_request.json()['data']['result']

            rate = 0
            for result in results:
                rate = float('{value[1]}'.format(**result))
                if rate is None or math.isnan(rate):
                    rate = 0

        except requests.exceptions.RequestException as e:
            rate = 0
        # print("fn cpu request: " + app + ": ", rate)

        return rate

        # for result in results:
        #     return float('{value[1]}'.format(**result))

    # Returned in MB
    def get_fn_memory_request(self, app):

        try:
            fn_memory_request = requests.get(PROMETHEUS + '/api/v1/query', params={
                'query': 'kube_pod_container_resource_requests{ namespace="default", container="' + app + '" , resource = '
                                                                                                          '"memory"}'})

            results = fn_memory_request.json()['data']['result']

            rate = 0
            for result in results:
                rate = float('{value[1]}'.format(**result))
                if rate is None or math.isnan(rate):
                    rate = 0

        except requests.exceptions.RequestException as e:
            rate = 0

        # print("Fn memory request: " + app + ": ", rate / 1000000)

        return rate / 1000000

    def get_fn_replica_no(self, app):

        try:
            fn_replica_no = requests.get(PROMETHEUS + '/api/v1/query', params={
                'query': 'kube_deployment_status_replicas_available{deployment = "' + app + '"}'})

            results = fn_replica_no.json()['data']['result']

            rate = 0
            for result in results:
                rate = float('{value[1]}'.format(**result))
                if rate is None or math.isnan(rate):
                    rate = 0

        except requests.exceptions.RequestException as e:
            rate = 0

        for p in pending_pod_list:
            if p.split('-')[0] == app:
                rate += 1

        return rate

        # for result in results:
        #     return float('{value[1]}'.format(**result))

    def scheduler(self, name, node, namespace="default"):
        target = client.V1ObjectReference()
        target.kind = "Node"
        target.apiVersion = "v1"
        target.name = node
        meta = client.V1ObjectMeta()
        meta.name = name
        body = client.V1Binding(target=target, metadata=meta)

        try:
            res = v1.create_namespaced_binding(namespace, body, _preload_content=False)
            if res:
                # print 'POD '+name+' scheduled and placed on '+node
                return True
        # return v1.create_namespaced_binding(namespace, body, _preload_content=False)
        except Exception as a:
            print("Exception when calling CoreV1Api->create_namespaced_binding: %s\n" % a)
            return False

    def get_node_cpu_requests(self):
        global node_cpu_requests_dict
        global node_cpu_requests_ratio_dict
        node_cpu_requests_dict = {}
        node_cpu_requests_ratio_dict = {}
        for node in NODE_NAME:
            # print("Status of node: ", node)

            try:
                total_cpu_requests = requests.get(PROMETHEUS + '/api/v1/query', params={
                    'query': 'sum(kube_pod_container_resource_requests{node= "' + node + '", resource = "cpu"})'})

                results = total_cpu_requests.json()['data']['result']
                cpu_requests = 0
                for result in results:
                    cpu_requests = float('{value[1]}'.format(**result))
                    if cpu_requests is None or math.isnan(cpu_requests):
                        cpu_requests = 0

            except requests.exceptions.RequestException as e:
                cpu_requests = 0

            r = {node: cpu_requests}
            node_cpu_requests_dict.update(r)
            r1 = {node: round((cpu_requests / NODE_CPU_CAP[node]), 2)}
            node_cpu_requests_ratio_dict.update(r1)

        return node_cpu_requests_dict

    def filtered_unavail_action_list(self, fn):
        unavail_action_list = []

        for node in node_cpu_requests_dict:
            if (CPU_REQUESTS[fn]) > (NODE_CPU_CAP[node] - 0.5 - node_cpu_requests_dict[node]):
                unavail_action_list.append(NODE_MAP_NAME_INDEX[node])

        return unavail_action_list

    #     def available_actions(self, fn):
    #         filtered_node_list = []
    #         for node in NODE_NAME:
    #             print("Status of node: ", node)
    #
    #             try:
    #                 total_cpu_requests = requests.get(PROMETHEUS + '/api/v1/query', params={
    #                     'query': 'sum(kube_pod_container_resource_requests{node= "' + node + '", resource = "cpu"})'})
    #
    #                 results = total_cpu_requests.json()['data']['result']
    #                 cpu_requests = 0
    #                 for result in results:
    #                     cpu_requests = float('{value[1]}'.format(**result))
    #                     if cpu_requests is None or math.isnan(cpu_requests):
    #                         cpu_requests = 0
    #
    #             except requests.exceptions.RequestException as e:
    #                 cpu_requests = 0
    #
    #   #*******************************************
    #
    #             # for result in results:
    #             #     total_cpu_requests = '{value[1]}'.format(**result)
    #             # print("total_cpu_requests of " + node + ": ", cpu_requests)
    #             ##
    #
    #             # try:
    #             #     total_memory_requests = requests.get(PROMETHEUS + '/api/v1/query', params={
    #             #         'query': 'sum(kube_pod_container_resource_requests{node= "' + node + '", resource = "memory"})'})
    #             #
    #             #     results = total_memory_requests.json()['data']['result']
    #             #
    #             #     memory_requests = 0
    #             #     for result in results:
    #             #         memory_requests = float('{value[1]}'.format(**result))
    #             #         if memory_requests is None or math.isnan(memory_requests):
    #             #             memory_requests = 0
    #             #
    #             # except requests.exceptions.RequestException as e:
    #             #     memory_requests = 0
    #
    #             # for result in results:
    #             #     total_memory_requests = '{value[1]}'.format(**result)
    #             #     if total_memory_requests is None:
    #             #         total_memory_requests = 0
    #
    #             # print("total_memory_requests of " + node + ": ", memory_requests)
    #             ##
    #
    #             # try:
    #             #     total_cpu_cores = requests.get(PROMETHEUS + '/api/v1/query', params={
    #             #         'query': 'node:node_num_cpu:sum{node="' + node + '"}'})
    #             #
    #             #     results = total_cpu_cores.json()['data']['result']
    #             #
    #             #     cpu_cores = 0
    #             #     for result in results:
    #             #         cpu_cores = float('{value[1]}'.format(**result))
    #             #         if cpu_cores is None or math.isnan(cpu_cores):
    #             #             cpu_cores = 0
    #             #
    #             # except requests.exceptions.RequestException as e:
    #             #     cpu_cores = 0
    #
    #             # for result in results:
    #             #     total_cpu_cores = '{value[1]}'.format(**result)
    #
    #             # print("total_cpu_cores of " + node + ": ", cpu_cores)
    #             ##
    #
    #             # try:
    #             #     total_available_memory = requests.get(PROMETHEUS + '/api/v1/query', params={
    #             #         'query': 'node_memory_MemTotal_bytes{instance="' + NODE_MAP_NAME_IP[node] + '"}'})
    #             #
    #             #     results = total_available_memory.json()['data']['result']
    #             #
    #             #     available_memory = 0
    #             #     for result in results:
    #             #         available_memory = float('{value[1]}'.format(**result))
    #             #         if available_memory is None or math.isnan(available_memory):
    #             #             available_memory = 0
    #             #
    #             # except requests.exceptions.RequestException as e:
    #             #     available_memory = 0
    #
    #             # for result in results:
    #             #     total_available_memory = '{value[1]}'.format(**result)
    #
    #             # print("total_available_memory of " + node + ": ", available_memory)
    # #************************************************************
    #             cpu_cores = 4
    #             # if (CPU_REQUESTS[fn] < (cpu_cores - cpu_requests) and self.get_fn_memory_request(fn) < (
    #             #         available_memory - memory_requests - 50000000)):
    #             #     FILTERED_NODE_LIST.append(NODE_MAP_NAME_IP[node])
    #
    #             if CPU_REQUESTS[fn] < (cpu_cores - cpu_requests):
    #                 filtered_node_list.append(NODE_MAP_NAME_INDEX[node])
    #
    #         return filtered_node_list

    def get_current_state(self, fn, row, ep, sch_data, wb):
        x = 0
        state = np.zeros(77)
        state2 = np.zeros(77)
        print(state.shape)
        global active_node_list
        active_node_list = []
        global active_nodes
        active_nodes = 0
        global active_fns
        global checked_fns
        global cpu_util_dict
        cpu_util_dict = {}
        active_fns = {"float": 0, "load": 0, "matrix": 0, "primary": 0, "dd": 0, "imageprocess": 0, "imageresize": 0,
                      "todoadditem": 0, "tododelitem": 0, "todogitem": 0, "todolistitem": 0, "todoupitem": 0,
                      "uploadimage": 0, "uploadvideo": 0, "video": 0}
        checked_fns = {"float": False, "load": False, "matrix": False, "primary": False, "dd": False,
                       "imageprocess": False, "imageresize": False, "todoadditem": False, "tododelitem": False,
                       "todogitem": False, "todolistitem": False, "todoupitem": False, "uploadimage": False,
                       "uploadvideo": False, "video": False}
        host_cpu_requests = self.get_node_cpu_requests()
        # print("Node cpu requests at thread %s: %s" % (str(row-1), str(host_cpu_requests)))
        for node in NODE_IP:
            # print("Status of node: " + NODE_MAP_IP_NAME[node])
            host_cpu_util = round(get_host_cpu_util(node) / 100, 2)
            # print("host_cpu_util of " + NODE_MAP_IP_NAME[node] + ": ", host_cpu_util)
            # we receive 0 usage when host is fully utilized and busy
            if host_cpu_util == 0:
                host_cpu_util = 1
            state[x] = host_cpu_util
            # state2[x] = host_cpu_util
            # if host_cpu_util < 0.5:
            #     self.state[x] = 0
            # if host_cpu_util >= 0.5 and host_cpu_util < 0.8:
            #     state[x] = 1
            # if host_cpu_util >= 0.8:
            #     state[x] = 2

            host_memory_util = self.get_host_memory_util(node)
            # print("host_memory_util of " + NODE_MAP_IP_NAME[node] + ": ", host_memory_util)
            state[x + 1] = round(host_memory_util / 100, 2)
            # state2[x + 1] = host_memory_util
            # if host_memory_util < 0.5:
            #     state[x + 1] = 0
            # if host_memory_util >= 0.5 and host_memory_util < 0.8:
            #     state[x + 1] = 1
            # if host_memory_util >= 0.8:
            #     state[x + 1] = 2

            # removed
            # host_bw_util = self.get_host_bw_util(node)
            # # print("host_bw_util of " + NODE_MAP_IP_NAME[node] + ": ", host_bw_util)
            # state[x + 2] = round(host_bw_util / 2000, 3)

            ############
            # state2[x + 2] = host_bw_util
            # if host_bw_util < 10:
            #     state[x + 2] = 0
            # if host_bw_util >= 10 and host_bw_util < 50:
            #     state[x + 2] = 1
            # if host_bw_util >= 50:
            #     state[x + 2] = 2

            # removed
            # host_diskio_time = self.get_host_diskio_util(node)
            # # print("host_diskio_time of " + NODE_MAP_IP_NAME[node] + ": ", host_diskio_time)
            # state[x + 3] = round(host_diskio_time / 1000, 3)

            ###########

            # state2[x + 3] = host_diskio_time
            # if host_diskio_time < 1:
            #     state[x + 3] = 0
            # if host_diskio_time >= 1 and host_diskio_time < 10:
            #     state[x + 3] = 1
            # if host_diskio_time >= 10:
            #     state[x + 3] = 2

            node_used_cpu_requests = host_cpu_requests[NODE_MAP_IP_NAME[node]]
            if NODE_MAP_IP_NAME[node] == 'node-9' or NODE_MAP_IP_NAME[node] == 'node-10' or NODE_MAP_IP_NAME[
                node] == 'node-15' or NODE_MAP_IP_NAME[node] == 'node-16' or NODE_MAP_IP_NAME[node] == 'node-17' or \
                    NODE_MAP_IP_NAME[node] == 'node-18':
                state[x + 2] = round(node_used_cpu_requests / 2, 3)
                # adding node cpu capacity
                state[x + 3] = (2 / 8)
                # print("this is node %s and node capacity is %s cpus" % ((str(NODE_MAP_IP_NAME[node])), str(state[x+3])))
            elif NODE_MAP_IP_NAME[node] == 'node-11' or NODE_MAP_IP_NAME[node] == 'node-12' or NODE_MAP_IP_NAME[
                node] == 'node-13' or NODE_MAP_IP_NAME[node] == 'node-14':
                state[x + 2] = round(node_used_cpu_requests / 8, 3)
                state[x + 3] = (8 / 8)
                # print("this is node %s and node capacity is %s cpus" % ((str(NODE_MAP_IP_NAME[node])), str(state[x+3])))

            else:
                state[x + 2] = round(node_used_cpu_requests / 4, 3)
                state[x + 3] = (4 / 8)
            # state2[x + 4] = node_used_cpu_requests
            #     print("this is node %s and node capacity is %s cpus" % ((str(NODE_MAP_IP_NAME[node])), str(state[x+3])))

            node_status, fn_replicas = get_node_status(node, fn)
            state[x + 4] = node_status
            # replicas of the scheduling fn on the vm
            state[x + 5] = fn_replicas / max_replicas
            # state2[x + 5] = node_status

            x = x + 6

        # fn_type = round(FN_NAME_INDEX[fn] / 5, 1)
        # fn_type = round(6 / 6, 1)
        # # print("fn to schedule is " + fn)
        # state[x] = fn_type

        fn_CPU = round(CPU_REQUESTS[fn], 2)
        # print("fn to schedule is " + fn)
        state[x] = fn_CPU
        # state2[x] = fn_type
        # if fn_cpu_request < 0.3:
        #     self.state[x] = 0
        # if fn_cpu_request >= 0.3 and fn_cpu_request < 0.5:
        #     state[x] = 1
        # if fn_cpu_request >= 0.5:
        #     state[x] = 2

        # fn_memory_request = self.get_fn_memory_request(fn)
        # print("fn_memory_request of " + fn + ": ", fn_memory_request)
        # self.state[x + 1] = 0
        # # if fn_memory_request < 100:
        # #     state[x + 1] = 0
        # # if fn_memory_request >= 100 and fn_memory_request < 300:
        # #     state[x + 1] = 1
        # # if fn_memory_request >= 300:
        # #     state[x + 1] = 2
        #
        # NW_BW_USAGE = POD_NW_BW_USAGE[fn]
        # print("fn_NW_BW_USAGE of " + fn + ": ", NW_BW_USAGE)
        # self.state[x + 2] = 0
        # # if NW_BW_USAGE < 10:
        # #     state[x + 2] = 0
        # # if NW_BW_USAGE >= 10 and NW_BW_USAGE < 50:
        # #     state[x + 2] = 1
        # # if NW_BW_USAGE >= 50:
        # #     state[x + 2] = 2

        # state2[x + 1] = fn_arrival_rate
        # if fn_arrival_rate < 20:
        #     state[x + 3] = 0
        # if fn_arrival_rate >= 20 and fn_arrival_rate < 50:
        #     state[x + 3] = 1
        # if fn_arrival_rate >= 50 and fn_arrival_rate < 100:
        #     state[x + 3] = 2
        # if fn_arrival_rate >= 100 and fn_arrival_rate < 150:
        #     state[x + 3] = 3
        # if fn_arrival_rate >= 150:
        #     state[x + 3] = 4

        # state[x + 1] = round((self.get_fn_replica_no(fn) - 1) / 4, 1)

        fn_arrivals_array = get_fn_arrival_rate()
        fn_arrival_rate = round(fn_arrivals_array[fn] / max_arr_rate, 2)
        print("fn %s has an arrival rate of %s" % (fn, str(fn_arrival_rate)))

        if fn_arrival_rate == 0 or fn_arrival_rate <= 0.1:
            global no_arrivals
            no_arrivals = True
        else:
            no_arrivals = False
        # print("fn_arrival_rate of " + fn + ": ", fn_arrival_rate)
        # state[x + 2] = fn_arrival_rate

        # request concurrency on an instance normalized(instead of arrival rate and no of replicas as above commented. Dividing by 20 which is the max value that can be there, with arrival rate at 20 and 1 replica)
        state[x + 1] = round((fn_arrivals_array[fn] / self.get_fn_replica_no(fn)) / max_arr_rate, 2)
        # average fn latency in the cluster normalized
        state[x + 2] = round(get_app_latency(fn) / 30, 2)

        y = x + 3
        for app in FN_ARRAY:
            if app == fn:
                continue
            else:
                state[y] = round(fn_arrivals_array[app] / max_arr_rate, 2)
                y += 1
        # state2[x + 2] = self.get_fn_replica_no(fn)
        # print("fn_replica_no of " + fn + ": ", state[x + 2])
        # print("Current state identified: \n")
        # print(state)
        # sch_data.write(row, 4, np.array_str(state2))
        # wb.save("Episodic_Data" + str(ep) + ".csv")

        return state

    # def remember(self, state, action, reward, next_state, done):
    #     memory.append((state, action, reward, next_state, done))
    #     if len(memory) > self.train_start:
    #         if self.epsilon > self.epsilon_min:
    #             self.epsilon *= self.epsilon_decay
    #             print("changed Epsilon: " + str(self.epsilon))

    def keys_exists(self, element, *keys):
        '''
        Check if *keys (nested) exists in `element` (dict).
        '''
        # if not isinstance(element, dict):
        #     raise AttributeError('keys_exists() expects dict as first argument.')
        # if len(keys) == 0:
        #     raise AttributeError('keys_exists() expects at least two arguments, one given.')

        _element = element
        for key in keys:
            try:
                _element = _element[key]
            except KeyError:
                return False
        return True

    def act(self, state, step, ep, sch_data, wb, app):
        global mode
        if mode == "agent":
            global epsilon
            discarded_action_list = self.filtered_unavail_action_list(app)
            # print("Discarded actions for thread %s: %s" % (str(step-1), str(discarded_action_list)))
            if np.random.random() <= epsilon:
                exist_count = 1
                while exist_count > 0:
                    sel_action = random.randrange(self.action_size)
                    exist_count = discarded_action_list.count(sel_action)

                action_t = "Random"
                # sch_data.write(step, 9, "Random")
                # wb.save("Episodic_Data" + str(ep) + ".csv")
                # print("Random action selected: ", sel_action)
                return sel_action, action_t
            else:
                predict_q = self.main_network.predict(state)
                # print("Q values for the thread %s from network: %s" % (step-1, str(predict_q[0])))
                action_q_values = predict_q[0]

                for n in discarded_action_list:
                    action_q_values[n] = action_q_values[n] - maxsize

                # print("adjusted action_q_values for thread %s: %s" % (step-1, str(action_q_values)))

                sel_action = np.argmax(action_q_values)

                # print("Selected action for thread %s: %s" % (step - 1, str(sel_action)))
                action_t = "Network"
                # sch_data.write(step, 9, "Network")
                # wb.save("Episodic_Data" + str(ep) + ".csv")
                # print("Action selected from network: ", sel_action)
                return sel_action, action_t

        else:
            node_selected = False
            # Random scheduling
            if mode == "random":
                while not node_selected:
                    node = random.randrange(10)
                    if (CPU_REQUESTS[app]) <= (NODE_CPU_CAP[NODE_MAP_INDEX_NAME[node]] - 0.5 - node_cpu_requests_dict[
                        NODE_MAP_INDEX_NAME[node]]):
                        node_selected = True
                action_t = "random"


            # cluster scheduling
            elif mode == "cluster":
                print("mode is cluster")
                global d
                try:
                    fn_arrival_rate = requests.get(PROMETHEUS + '/api/v1/query', params={
                        'query': 'sum(increase(function_calls_total{function = "' + app + '" }[10s]))'})

                    results = fn_arrival_rate.json()['data']['result']

                    fn_requests = 0
                    for result in results:
                        fn_requests = float('{value[1]}'.format(**result))
                        if fn_requests is None or math.isnan(fn_requests):
                            fn_requests = 0

                except requests.exceptions.RequestException as e:
                    fn_requests = 0

                # if d[app][round(fn_requests)] in d:
                #     fn_cluster = d[app][round(fn_requests)]
                #
                # else:
                #     fn_cluster = 0

                key_there = self.keys_exists(d, app, round(fn_requests/10))
                if key_there:
                    fn_cluster = d[app][round(fn_requests/10)]
                else:
                    fn_cluster = 0

                print("fn is %s and AR is %s and cluster is %s" % (app, str(round(fn_requests/10)), str(fn_cluster)))

                for x in range(10):
                    node = x
                    if (CPU_REQUESTS[app]) <= (NODE_CPU_CAP[NODE_MAP_INDEX_NAME[node]] - 0.5 - node_cpu_requests_dict[
                        NODE_MAP_INDEX_NAME[node]]):

                        pod_list = []
                        node_clusters = []
                        try:
                            running_pods = requests.get(PROMETHEUS + '/api/v1/query', params={
                                'query': 'kube_pod_info{namespace="default", node="' + NODE_MAP_INDEX_NAME[
                                    node] + '"}'})

                            pods = running_pods.json()['data']['result']

                            # print("%s: Pods in node %s are %s" % (datetime.now(), node, str(pods)))

                            for pod in pods:
                                pod_data = {"fn": "", "ar": 0}
                                fn_name = (pod['metric']['pod']).split('-')[0]

                                if fn_name == "nginx" or fn_name == "test" or fn_name == "my":
                                    continue

                                try:
                                    fn_arrival_rate = requests.get(PROMETHEUS + '/api/v1/query', params={
                                        'query': 'sum(increase(function_calls_total{function = "' + fn_name + '" }[10s]))'})

                                    results = fn_arrival_rate.json()['data']['result']

                                    fn_requests = 0
                                    for result in results:
                                        fn_requests = float('{value[1]}'.format(**result))
                                        if fn_requests is None or math.isnan(fn_requests):
                                            fn_requests = 0

                                except requests.exceptions.RequestException as e:
                                    fn_requests = 0

                                pod_data['fn'] = fn_name
                                pod_data['ar'] = round(fn_requests/10)
                                pod_list.append(pod_data)

                        except requests.exceptions.RequestException as e:
                            pass

                        for x in range(len(pod_list)):
                            done = False
                            for f in d:
                                if pod_list[x]['fn'] == f:
                                    for rate in d[f]:
                                        if pod_list[x]['ar'] == rate:
                                            node_clusters.append(d[f][rate])
                                            done = True
                                            break
                                    if not done:
                                        node_clusters.append(0)
                                        done = True
                                if done:
                                    break

                        if fn_cluster not in node_clusters:
                            # # print("fn cluster %s is not in node clusters %s, so selecting it" % (
                            # str(fn_cluster), str(node_clusters)))
                            node_selected = True
                            break
                        # else:
                        #     print("fn cluster %s is in node clusters %s, so NOT selecting it" % (
                        #         str(fn_cluster), str(node_clusters)))

                if not node_selected:
                    node = random.randrange(10)
                    print("random node selected")
                else:
                    print("Actual node selected by clustering")
                action_t = "cluster"

                #             # print("%s: Fn requests for fn %s are %s" % (datetime.now(), fn_name, str(fn_requests)))
                #
                #             # for result in results:
                #             #     fn_requests = float('{value[1]}'.format(**result))
                #             checked_fns[fn_name] = True
                #             if fn_requests > 0:
                #                 active_nodes += 1
                #                 active_node_list.append(node)
                #                 node_active = 1
                #                 # print("Active node>>> ", node)
                #                 active_fns[fn_name] = 1
                #                 break
                #         else:
                #             if active_fns[fn_name] > 0:
                #                 active_nodes += 1
                #                 active_node_list.append(node)
                #                 node_active = 1
                #                 # print("Node %s is Active since fn %s is in active list " % (str(node), fn_name))
                #                 break
                #
                # node_selected = True


            # RR scheduling
            elif mode == "RR":
                global rr_node_no
                while not node_selected:
                    node = rr_node_no
                    if (CPU_REQUESTS[app]) <= (NODE_CPU_CAP[NODE_MAP_INDEX_NAME[node]] - 0.5 - node_cpu_requests_dict[
                        NODE_MAP_INDEX_NAME[node]]):
                        node_selected = True
                    else:
                        if rr_node_no == 9:
                            rr_node_no = 0
                        else:
                            rr_node_no += 1

                if rr_node_no == 9:
                    rr_node_no = 0
                else:
                    rr_node_no += 1
                action_t = "RR"

            elif mode == "BPFF":
                for x in range(10):
                    if (CPU_REQUESTS[app]) <= (NODE_CPU_CAP[NODE_MAP_INDEX_NAME[x]] - 0.5 - node_cpu_requests_dict[
                        NODE_MAP_INDEX_NAME[x]]):
                        node = x
                        # node_selected = True
                        break
                action_t = "BPFF"

            elif mode == "BPBF":
                least_free_cpu = 8
                for x in range(10):
                    free_cpu = NODE_CPU_CAP[NODE_MAP_INDEX_NAME[x]] - 0.5 - node_cpu_requests_dict[
                        NODE_MAP_INDEX_NAME[x]]
                    if (CPU_REQUESTS[app]) <= free_cpu:
                        if free_cpu < least_free_cpu:
                            node = x
                            least_free_cpu = free_cpu
                        # node_selected = True
                action_t = "BPBF"

            elif mode == "LC":
                least_connections = 1000
                for x in range(10):
                    pod_no = get_node_pod_no(NODE_MAP_INDEX_NAME[x])
                    free_cpu = NODE_CPU_CAP[NODE_MAP_INDEX_NAME[x]] - 0.5 - node_cpu_requests_dict[
                        NODE_MAP_INDEX_NAME[x]]
                    if (CPU_REQUESTS[app]) <= free_cpu:
                        if pod_no < least_connections:
                            node = x
                            least_connections = pod_no
                        # node_selected = True
                action_t = "LC"

            elif mode == "score":
                # print("at score")
                node_cpu_util_dict = {}
                node_mem_util_dict = {}
                node_status_dict = {}
                node_score = {0: 0, 1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0}
                for y in range(10):
                    util = get_host_cpu_util(NODE_MAP_NAME_IP[NODE_MAP_INDEX_NAME[y]])
                    r = {y: util}
                    node_cpu_util_dict.update(r)
                sorted_cpu = sorted(node_cpu_util_dict.items(), key=lambda kv: kv[1])
                sorted_dic_cpu = dict(sorted_cpu)

                for y in range(10):
                    util = self.get_host_memory_util(NODE_MAP_NAME_IP[NODE_MAP_INDEX_NAME[y]])
                    r = {y: util}
                    node_mem_util_dict.update(r)
                sorted_mem = sorted(node_mem_util_dict.items(), key=lambda kv: kv[1])
                sorted_dic_mem = dict(sorted_mem)

                sorted_req = sorted(node_cpu_requests_dict.items(), key=lambda kv: kv[1])
                sorted_dic_req = dict(sorted_req)

                i = 1
                for key in sorted_dic_cpu.keys():
                    sorted_dic_cpu[key] = i
                    i = i + 1

                i = 1
                for key in sorted_dic_mem.keys():
                    sorted_dic_mem[key] = i
                    i = i + 1

                i = 1
                for key in sorted_dic_req.keys():
                    sorted_dic_req[key] = i
                    i = i + 1

                i = 0

                for x in range(10):
                    exist_count = active_node_list.count(NODE_MAP_INDEX_NAME[x])
                    if exist_count > 0:
                        status = 1
                    else:
                        status = 0
                    r = {x: status}
                    node_status_dict.update(r)

                for z in range(10):
                    node_score[z] = sorted_dic_cpu[z] + sorted_dic_mem[z] + sorted_dic_req[NODE_MAP_INDEX_NAME[z]] + \
                                    node_status_dict[z]

                sorted_score = sorted(node_score.items(), key=lambda kv: kv[1], reverse=True)
                sorted_dic_score = dict(sorted_score)

                for key in sorted_dic_score.keys():
                    if (CPU_REQUESTS[app]) <= (NODE_CPU_CAP[NODE_MAP_INDEX_NAME[key]] - 0.5 - node_cpu_requests_dict[
                        NODE_MAP_INDEX_NAME[key]]):
                        node = key
                        # node_selected = True
                        break
                action_t = "score"

            elif mode == "STCA-H":
                # print("at score")
                node_cpu_util_dict = {}
                node_mem_util_dict = {}
                node_status_dict = {}
                node_score = {0: 0, 1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0}
                for y in range(10):
                    util = get_host_cpu_util(NODE_MAP_NAME_IP[NODE_MAP_INDEX_NAME[y]])
                    r = {y: util}
                    node_cpu_util_dict.update(r)
                sorted_cpu = sorted(node_cpu_util_dict.items(), key=lambda kv: kv[1])
                sorted_dic_cpu = dict(sorted_cpu)

                for y in range(10):
                    util = self.get_host_memory_util(NODE_MAP_NAME_IP[NODE_MAP_INDEX_NAME[y]])
                    r = {y: util}
                    node_mem_util_dict.update(r)
                sorted_mem = sorted(node_mem_util_dict.items(), key=lambda kv: kv[1])
                sorted_dic_mem = dict(sorted_mem)

                sorted_req = sorted(node_cpu_requests_ratio_dict.items(), key=lambda kv: kv[1])
                sorted_dic_req = dict(sorted_req)

                i = 1
                for key in sorted_dic_cpu.keys():
                    sorted_dic_cpu[key] = i
                    i = i + 1

                i = 1
                for key in sorted_dic_mem.keys():
                    sorted_dic_mem[key] = i
                    i = i + 1

                i = 1
                for key in sorted_dic_req.keys():
                    sorted_dic_req[key] = i
                    i = i + 1

                i = 0

                for x in range(10):
                    exist_count = active_node_list.count(NODE_MAP_INDEX_NAME[x])
                    if exist_count > 0:
                        status = 1
                    else:
                        status = 0
                    r = {x: status}
                    node_status_dict.update(r)

                for z in range(10):
                    node_score[z] = sorted_dic_cpu[z] + sorted_dic_mem[z] + sorted_dic_req[NODE_MAP_INDEX_NAME[z]] + \
                                    node_status_dict[z]

                sorted_score = sorted(node_score.items(), key=lambda kv: kv[1], reverse=True)
                sorted_dic_score = dict(sorted_score)

                for key in sorted_dic_score.keys():
                    if (CPU_REQUESTS[app]) <= (NODE_CPU_CAP[NODE_MAP_INDEX_NAME[key]] - 0.5 - node_cpu_requests_dict[
                        NODE_MAP_INDEX_NAME[key]]):
                        node = key
                        # node_selected = True
                        break
                action_t = "STCA-H"

            elif mode == "STCA-L":
                # print("at score")
                node_cpu_util_dict = {}
                node_mem_util_dict = {}
                node_status_dict = {}
                node_score = {0: 0, 1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0}
                for y in range(10):
                    util = get_host_cpu_util(NODE_MAP_NAME_IP[NODE_MAP_INDEX_NAME[y]])
                    r = {y: util}
                    node_cpu_util_dict.update(r)
                sorted_cpu = sorted(node_cpu_util_dict.items(), key=lambda kv: kv[1])
                sorted_dic_cpu = dict(sorted_cpu)

                for y in range(10):
                    util = self.get_host_memory_util(NODE_MAP_NAME_IP[NODE_MAP_INDEX_NAME[y]])
                    r = {y: util}
                    node_mem_util_dict.update(r)
                sorted_mem = sorted(node_mem_util_dict.items(), key=lambda kv: kv[1])
                sorted_dic_mem = dict(sorted_mem)

                sorted_req = sorted(node_cpu_requests_ratio_dict.items(), key=lambda kv: kv[1])
                sorted_dic_req = dict(sorted_req)

                i = 1
                for key in sorted_dic_cpu.keys():
                    sorted_dic_cpu[key] = i
                    i = i + 1

                i = 1
                for key in sorted_dic_mem.keys():
                    sorted_dic_mem[key] = i
                    i = i + 1

                i = 1
                for key in sorted_dic_req.keys():
                    sorted_dic_req[key] = i
                    i = i + 1

                i = 0

                for x in range(10):
                    exist_count = active_node_list.count(NODE_MAP_INDEX_NAME[x])
                    if exist_count > 0:
                        status = 1
                    else:
                        status = 0
                    r = {x: status}
                    node_status_dict.update(r)

                for z in range(10):
                    node_score[z] = sorted_dic_cpu[z] + sorted_dic_mem[z] + sorted_dic_req[NODE_MAP_INDEX_NAME[z]] + \
                                    node_status_dict[z]

                sorted_score = sorted(node_score.items(), key=lambda kv: kv[1])
                sorted_dic_score = dict(sorted_score)

                for key in sorted_dic_score.keys():
                    if (CPU_REQUESTS[app]) <= (NODE_CPU_CAP[NODE_MAP_INDEX_NAME[key]] - 0.5 - node_cpu_requests_dict[
                        NODE_MAP_INDEX_NAME[key]]):
                        node = key
                        # node_selected = True
                        break
                action_t = "STCA-L"

            # elif mode == "DTCA":
            #     avg_fn_lat = round(get_app_latency(app), 2)
            #     if avg_fn_lat > 1.25:
            #         value = 0
            #         for y in range(20):
            #             util = round(get_host_cpu_util(NODE_MAP_NAME_IP[NODE_MAP_INDEX_NAME[y]]) / 100, 2)
            #             cap = NODE_CPU_CAP[NODE_MAP_INDEX_NAME[y]]
            #             cost_rate = NODE_HR_RATE[NODE_MAP_INDEX_NAME[y]]
            #             n_score = round(((1 - util) * cap / cost_rate), 2)
            #             if n_score > value:
            #                 node = y
            #                 value = n_score
            #
            #     if avg_fn_lat <= 1.25:
            #         value = 0
            #         for y in range(20):
            #             util = round(get_host_cpu_util(NODE_MAP_NAME_IP[NODE_MAP_INDEX_NAME[y]]) / 100, 2)
            #             cap = NODE_CPU_CAP[NODE_MAP_INDEX_NAME[y]]
            #             cost_rate = NODE_HR_RATE[NODE_MAP_INDEX_NAME[y]]
            #             n_score = round((util * cap / cost_rate), 2)
            #             if n_score > value:
            #                 node = y
            #                 value = n_score
            #
            #     action_t = "DTCA"

            elif mode == "DTCA":
                avg_fn_lat = round(get_app_latency(app), 2)
                if avg_fn_lat > 1.25:
                    node_cpu_util_dict = {}
                    node_mem_util_dict = {}
                    node_status_dict = {}
                    node_score = {0: 0, 1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0}
                    for y in range(10):
                        util = get_host_cpu_util(NODE_MAP_NAME_IP[NODE_MAP_INDEX_NAME[y]])
                        r = {y: util}
                        node_cpu_util_dict.update(r)
                    sorted_cpu = sorted(node_cpu_util_dict.items(), key=lambda kv: kv[1])
                    sorted_dic_cpu = dict(sorted_cpu)

                    for y in range(10):
                        util = self.get_host_memory_util(NODE_MAP_NAME_IP[NODE_MAP_INDEX_NAME[y]])
                        r = {y: util}
                        node_mem_util_dict.update(r)
                    sorted_mem = sorted(node_mem_util_dict.items(), key=lambda kv: kv[1])
                    sorted_dic_mem = dict(sorted_mem)

                    sorted_req = sorted(node_cpu_requests_ratio_dict.items(), key=lambda kv: kv[1])
                    sorted_dic_req = dict(sorted_req)

                    i = 1
                    for key in sorted_dic_cpu.keys():
                        sorted_dic_cpu[key] = i
                        i = i + 1

                    i = 1
                    for key in sorted_dic_mem.keys():
                        sorted_dic_mem[key] = i
                        i = i + 1

                    i = 1
                    for key in sorted_dic_req.keys():
                        sorted_dic_req[key] = i
                        i = i + 1

                    i = 0

                    for x in range(10):
                        exist_count = active_node_list.count(NODE_MAP_INDEX_NAME[x])
                        if exist_count > 0:
                            status = 1
                        else:
                            status = 0
                        r = {x: status}
                        node_status_dict.update(r)

                    for z in range(10):
                        node_score[z] = sorted_dic_cpu[z] + sorted_dic_mem[z] + sorted_dic_req[NODE_MAP_INDEX_NAME[z]] + \
                                        node_status_dict[z]

                    sorted_score = sorted(node_score.items(), key=lambda kv: kv[1])
                    sorted_dic_score = dict(sorted_score)

                    for key in sorted_dic_score.keys():
                        if (CPU_REQUESTS[app]) <= (
                                NODE_CPU_CAP[NODE_MAP_INDEX_NAME[key]] - 0.5 - node_cpu_requests_dict[
                            NODE_MAP_INDEX_NAME[key]]):
                            node = key
                            # node_selected = True
                            break

                elif avg_fn_lat <= 1.25:
                    node_cpu_util_dict = {}
                    node_mem_util_dict = {}
                    node_status_dict = {}
                    node_score = {0: 0, 1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0}
                    for y in range(10):
                        util = get_host_cpu_util(NODE_MAP_NAME_IP[NODE_MAP_INDEX_NAME[y]])
                        r = {y: util}
                        node_cpu_util_dict.update(r)
                    sorted_cpu = sorted(node_cpu_util_dict.items(), key=lambda kv: kv[1])
                    sorted_dic_cpu = dict(sorted_cpu)

                    for y in range(10):
                        util = self.get_host_memory_util(NODE_MAP_NAME_IP[NODE_MAP_INDEX_NAME[y]])
                        r = {y: util}
                        node_mem_util_dict.update(r)
                    sorted_mem = sorted(node_mem_util_dict.items(), key=lambda kv: kv[1])
                    sorted_dic_mem = dict(sorted_mem)

                    sorted_req = sorted(node_cpu_requests_ratio_dict.items(), key=lambda kv: kv[1])
                    sorted_dic_req = dict(sorted_req)

                    i = 1
                    for key in sorted_dic_cpu.keys():
                        sorted_dic_cpu[key] = i
                        i = i + 1

                    i = 1
                    for key in sorted_dic_mem.keys():
                        sorted_dic_mem[key] = i
                        i = i + 1

                    i = 1
                    for key in sorted_dic_req.keys():
                        sorted_dic_req[key] = i
                        i = i + 1

                    i = 0

                    for x in range(10):
                        exist_count = active_node_list.count(NODE_MAP_INDEX_NAME[x])
                        if exist_count > 0:
                            status = 1
                        else:
                            status = 0
                        r = {x: status}
                        node_status_dict.update(r)

                    for z in range(10):
                        node_score[z] = sorted_dic_cpu[z] + sorted_dic_mem[z] + sorted_dic_req[NODE_MAP_INDEX_NAME[z]] + \
                                        node_status_dict[z]

                    sorted_score = sorted(node_score.items(), key=lambda kv: kv[1], reverse=True)
                    sorted_dic_score = dict(sorted_score)

                    for key in sorted_dic_score.keys():
                        if (CPU_REQUESTS[app]) <= (
                                NODE_CPU_CAP[NODE_MAP_INDEX_NAME[key]] - 0.5 - node_cpu_requests_dict[
                            NODE_MAP_INDEX_NAME[key]]):
                            node = key
                            # node_selected = True
                            break

                action_t = "DTCA"

            elif mode == "ENSURE":
                avg_fn_lat = round(get_app_latency(app), 2)
                if avg_fn_lat > 1.25:
                    for x in range(10):
                        if (CPU_REQUESTS[app]) <= (NODE_CPU_CAP[NODE_MAP_INDEX_NAME[x]] - 0.5 - node_cpu_requests_dict[
                            NODE_MAP_INDEX_NAME[x]]):
                            node_status, fn_replicas = get_node_status(NODE_MAP_NAME_IP[NODE_MAP_INDEX_NAME[x]], app)
                            if fn_replicas < 1:
                                node = x
                                # node_selected = True
                                break

                elif avg_fn_lat <= 1.25:
                    for x in range(10):
                        if (CPU_REQUESTS[app]) <= (NODE_CPU_CAP[NODE_MAP_INDEX_NAME[x]] - 0.5 - node_cpu_requests_dict[
                            NODE_MAP_INDEX_NAME[x]]):
                            node_status, fn_replicas = get_node_status(NODE_MAP_NAME_IP[NODE_MAP_INDEX_NAME[x]], app)
                            if fn_replicas < NODE_CPU_CAP[NODE_MAP_INDEX_NAME[x]]:
                                node = x
                                # node_selected = True
                                break
                        # node_selected = True
                action_t = "ENSURE"

            return node, action_t

    def update_target_network(self):
        self.target_network.set_weights(self.main_network.get_weights())

    def replay(self, terminal_state):
        if len(memory) >= train_start:
            # return
            # Randomly sample minibatch from the memory
            print("Memory replaying...")

            # global step_count_total
            # if step_count_total % 50 == 0:
            #     agent.tensorboard.step += 1
            #     print("Tensorboard step: " + str(agent.tensorboard.step))

            # minibatch = random.sample(memory, min(len(memory), self.batch_size))
            minibatch = random.sample(memory, self.batch_size)

            # Get current states from minibatch, then query NN model for Q values
            # np.reshape(current_state, [1, self.state_size])
            # current_states = np.array([transition[0] for transition in minibatch])
            # current_states = np.array([transition[0] for transition in minibatch])
            # current_qs_list = self.main_network.predict(current_states)
            #
            # # Get future states from minibatch, then query NN model for Q values
            # # When using target network, query it, otherwise main network should be queried
            # new_current_states = np.array([transition[3] for transition in minibatch])
            # future_qs_list = self.target_network.predict(new_current_states)
            #
            # X = []
            # y = []
            #
            # # Now we need to enumerate our batches
            # for index, (current_state, action, reward, new_current_state, done) in enumerate(minibatch):
            #
            #     # If not a terminal state, get new q from future states, otherwise set it to 0
            #     # almost like with Q Learning, but we use just part of equation here
            #     if not done:
            #         max_future_q = np.max(future_qs_list[index])
            #         new_q = reward + self.gamma * max_future_q
            #     else:
            #         new_q = reward
            #
            #     # Update Q value for given state
            #     current_qs = current_qs_list[index]
            #     current_qs[action] = new_q
            #
            #     # And append to our training data
            #     X.append(current_state)
            #     y.append(current_qs)
            #
            # # Fit on all samples as one batch, log only on terminal state
            # self.main_network.fit(np.array(X), np.array(y), batch_size=self.batch_size, verbose=0, shuffle=False,
            #                callbacks=[self.tensorboard] if terminal_state else None)

            # current_state = np.zeros((self.batch_size, self.state_size))
            # next_state = np.zeros((self.batch_size, self.state_size))
            # action, reward, done = [], [], []
            #
            # # do this before prediction
            # # for speedup, this could be done on the tensor level
            # # but easier to understand using a loop
            # for i in range(self.batch_size):
            #     current_state[i] = np.reshape(minibatch[i][0], [1, self.state_size])
            #     action.append(minibatch[i][1])
            #     reward.append(minibatch[i][2])
            #     next_state[i] = np.reshape(minibatch[i][3], [1, self.state_size])
            #     done.append(minibatch[i][4])
            #
            # # do batch prediction to save speed
            # target = self.main_network.predict(current_state)
            # target_next = self.target_network.predict(next_state)
            #
            # for i in range(self.batch_size):
            #     # correction on the Q value for the action used
            #     if done[i]:
            #         target[i][action[i]] = reward[i]
            #     else:
            #         # Standard - DQN
            #         # DQN chooses the max Q value among next actions
            #         # selection and evaluation of action is on the target Q Network
            #         # Q_max = max_a' Q_target(s', a')
            #         target[i][action[i]] = reward[i] + self.gamma * (np.amax(target_next[i]))
            #
            # # Train the Neural Network with batches
            # # self.model.fit(current_state, target, batch_size=self.batch_size, verbose=0)
            # self.main_network.fit(current_state, target, batch_size=self.batch_size, verbose=0, shuffle=False,
            #                       callbacks=[self.tensorboard] if terminal_state else None)

            # for state, action, reward, next_state, done in minibatch:
            #     if not done:
            #         target_q = (reward + self.gamma * np.max(
            #             self.target_network.predict(next_state)))
            #     else:
            #         target_q = reward
            #
            #     predict_q = self.main_network.predict(state)
            #     predict_q[0][action] = target_q
            #
            #     # Train the Neural Network with batches
            #     self.main_network.fit(state, predict_q, batch_size=self.batch_size, verbose=0, shuffle=False,
            #                           callbacks=[tensorboard_callback] if terminal_state else None)

            state = np.zeros((self.batch_size, self.state_size))
            # next_state = np.zeros((self.batch_size, self.state_size))
            next_state = np.zeros((self.batch_size, self.state_size))
            action, reward, done = [], [], []

            # do this before prediction
            # for speedup, this could be done on the tensor level
            # but easier to understand using a loop
            for i in range(self.batch_size):
                state[i] = minibatch[i][0]
                action.append(minibatch[i][1])
                reward.append(minibatch[i][2])
                next_state[i] = minibatch[i][3]
                done.append(minibatch[i][4])

            # do batch prediction to save speed - here target means a Q value.
            target = self.main_network.predict(state)
            target_next = self.target_network.predict(next_state)

            for i in range(self.batch_size):
                # correction on the Q value for the action used
                if done[i]:
                    target[i][action[i]] = reward[i]
                else:
                    # Standard - DQN
                    # DQN chooses the max Q value among next actions
                    # selection and evaluation of action is on the target Q Network
                    # Q_max = max_a' Q_target(s', a')
                    target[i][action[i]] = reward[i] + self.gamma * (np.amax(target_next[i]))

            self.main_network.fit(state, target, batch_size=self.batch_size, verbose=0, shuffle=False,
                                  callbacks=[self.tensorboard] if terminal_state else None)

        if terminal_state:
            agent.tensorboard.update_stats(Episodic_reward=episodic_reward)
            agent.tensorboard.update_stats(Function_Latency=function_latency)
            agent.tensorboard.update_stats(Total_VM_TIME_DIFF=total_vm_time_diff)
            agent.tensorboard.update_stats(Total_VM_COST_DIFF=total_vm_cost_diff)
            agent.tensorboard.update_stats(Active_node_penalty=act_node_penalty)
            # agent.tensorboard.update_stats(CPU_util_penalty=cpu_util_penalty)

        print("Tensorboard step: " + str(agent.tensorboard.step))

    def graphs(self, write_graph):
        if write_graph:
            agent.tensorboard.update_stats(Episodic_reward=episodic_reward)
            agent.tensorboard.update_stats(Function_Latency=function_latency)
            agent.tensorboard.update_stats(Total_VM_TIME_DIFF=total_vm_time_diff)
            agent.tensorboard.update_stats(Total_VM_COST_DIFF=total_vm_cost_diff)
            agent.tensorboard.update_stats(Active_node_penalty=act_node_penalty)
            # agent.tensorboard.update_stats(CPU_util_penalty=cpu_util_penalty)

    def load(self, name):
        self.main_network = load_model(name)

    def save(self, name):
        self.main_network.save(name)

    def run(self):
        # epsilon = 1
        print("Starting model training")
        wb_inf = Workbook()
        inference = wb_inf.add_sheet('Inf_time')
        inference.write(0, 0, 'Episode')
        inference.write(0, 1, 'Event_No')
        inference.write(0, 2, 'Inf_time')

        if mode == 'cluster':
            global d
            wb = openpyxl.load_workbook('data_out.xlsx')
            sh = wb['Sheet1']
            # 752
            for i in range(2, 752):
                fn = sh.cell(i, 2).value
                # print(fn)
                if fn in d:
                    ar = sh.cell(i, 3).value
                    d[fn][ar] = sh.cell(i, 8).value
                    # print(d[fn][ar])
                else:
                    d[fn] = {}
                    ar = sh.cell(i, 3).value
                    d[fn][ar] = sh.cell(i, 8).value
                    # print(d[fn][ar])

        for e in range(self.EPISODES):
            global training_end
            global episodic_reward
            episodic_reward = 0
            global function_latency
            function_latency = 0
            global total_vm_time_diff
            total_vm_time_diff = 0
            global total_vm_cost_diff
            total_vm_cost_diff = 0
            global cpu_util_penalty
            cpu_util_penalty = 0
            global act_node_penalty
            act_node_penalty = 0
            global epsilon
            destroyed_pod_list = []
            global initial_lat
            initial_lat = 0
            wb = Workbook()
            sch_data = wb.add_sheet('Scheduled')
            sch_data.write(0, 0, 'Time')
            sch_data.write(0, 1, 'Episode')
            sch_data.write(0, 2, 'Event_No')
            sch_data.write(0, 3, 'Pod')
            sch_data.write(0, 4, 'State')
            sch_data.write(0, 5, 'Action')
            sch_data.write(0, 6, 'Reward')
            sch_data.write(0, 7, 'Next_state')
            sch_data.write(0, 8, 'Done')
            sch_data.write(0, 9, 'Random or NOT')
            sch_data.write(0, 10, 'Pod_created_time')
            sch_data.write(0, 11, 'Fn_latency')
            sch_data.write(0, 12, 'Fn_failure_rate')
            sch_data.write(0, 13, 'CPU_penalty')
            sch_data.write(0, 14, 'Active_node_no')
            sch_data.write(0, 15, 'VM_up_time_cost')
            sch_data.write(0, 16, 'VM_up_time')

            ep_data = wb.add_sheet('Episodes')
            ep_data.write(0, 0, 'Time')
            ep_data.write(0, 1, 'Episode')
            ep_data.write(0, 2, 'Epsilon')
            ep_data.write(0, 3, 'Ep_reward')
            ep_data.write(0, 4, 'Avg_nodes')

            reward_q.queue.clear()
            state_q.queue.clear()
            action_q.queue.clear()
            done_q.queue.clear()
            next_state_q.queue.clear()
            threadID_q.queue.clear()
            global reward_dict
            reward_dict = {}
            z = 0
            done = False
            start_thread = False

            # state = self.env.reset()
            event = 0
            try:
                w = watch.Watch()
                for event in w.stream(v1.list_namespaced_pod, "default"):
                    # print("z value at the beginning: ", z)
                    #     if event == None:
                    #         break
                    #     if event['object'].status.phase == "Pending" and event['object'].status.conditions == None and event['object'].spec.scheduler_name == scheduler_name:
                    #             fn_to_schedule = event['object'].metadata.name.split('-')[0]
                    #             print(fn_to_schedule)
                    #
                    #
                    # print("no more pods to schedule")
                    #
                    #     # print(w.stream(v1.list_namespaced_pod, "default"))

                    if event['object'].status.phase == "Pending" and event['object'].status.conditions == None and \
                            event[
                                'object'].spec.scheduler_name == scheduler_name:
                        write_log = False

                        pod_false = False
                        for dest_pod in destroyed_pod_list:
                            if event['object'].metadata.name == dest_pod:
                                pod_false = True
                                print("Pod is in the destroyed pod list so not scheduling it")
                                break
                        if pod_false:
                            continue

                        if z == 0 and start_thread is False:
                            global stop_an_thread
                            print("starting vm info thread %s from main loop" % ("ep" + str(e)))
                            stop_an_thread = False
                            start_thread = True
                            act_n_thread = "ep" + str(e)
                            active_node_thread = epThread(act_n_thread, ep_data, wb, e + 1, e)
                            active_node_thread.start()

                        try:
                            start = datetime.now()
                            print("% *************************Start Timestamp: ", start)

                            # pod_false = False
                            # for dest_pod in destroyed_pod_list:
                            #     if event['object'].metadata.name == dest_pod:
                            #         pod_false = True
                            #         print("Pod is in the destroyed pod list so not scheduling it")
                            #         break
                            # if pod_false:
                            #     continue

                            fn_to_schedule = event['object'].metadata.name.split('-')[0]
                            if fn_to_schedule == "test":
                                # global stop_an_thread
                                stop_an_thread = True
                                training_end = True
                                done = True
                                done_q.put(done)
                                # reward_q.put(0)
                                current_state = np.zeros(77)
                                current_state = np.reshape(current_state, [1, self.state_size])
                                # state_q.put(current_state)
                                next_state_q.put(current_state)
                                next_state = current_state
                                res = self.scheduler(event['object'].metadata.name, "node-3")
                                while state_q.qsize() != 0 and reward_q.qsize() != 0 and action_q.qsize() != 0 and next_state_q.qsize() != 0 and done_q.qsize() != 0:
                                    print("Appending to buffer at episode end")
                                    # memory.append(
                                    #     (state_q.get(), action_q.get(), reward_q.get(), next_state_q.get(),
                                    #      done_q.get()))

                                    n_step_buffer.append(
                                        (state_q.get(), action_q.get(), reward_q.get(), next_state_q.get(),
                                         done_q.get()))
                                    print("n-step buffer length: ", len(n_step_buffer))

                                    if len(n_step_buffer) == n_step:  # fill the n-step buffer for the first translation
                                        # add a multi step transition
                                        reward, next_obs, done = get_n_step_info(n_step_buffer, gamma_nstep)
                                        obs, action = n_step_buffer[0][:2]

                                        memory.append((obs, action, reward, next_obs, done))
                                try:
                                    sch_data.write(z + 1, 0, datetime.now())
                                    sch_data.write(z + 1, 1, e)
                                    sch_data.write(z + 1, 2, z)
                                    sch_data.write(z + 1, 3, event['object'].metadata.name)
                                    # sch_data.write(z + 1, 4, np.array_str(current_state))
                                    sch_data.write(z + 1, 5, 'None')
                                    sch_data.write(z, 7, np.array_str(next_state))
                                    sch_data.write(z, 8, done)
                                    sch_data.write(z + 1, 9, "Training end")

                                    wb.save("episodic_data/Episodic_Data" + str(e) + ".csv")
                                    # print("Saved to Episodic data3")
                                except Exception as inst:
                                    print(type(inst))

                                break

                            if fn_to_schedule == "nginx":
                                # global stop_an_thread
                                stop_an_thread = True
                                print("NGINX pod")
                                done = True
                                done_q.put(done)
                                # reward_q.put(0)
                                current_state = np.zeros(77)
                                current_state = np.reshape(current_state, [1, self.state_size])
                                # state_q.put(current_state)
                                next_state_q.put(current_state)
                                next_state = current_state
                                res = self.scheduler(event['object'].metadata.name, "node-3")
                                # try:
                                #     sch_data.write(z, 7, np.array_str(next_state))
                                # except Exception as inst:
                                #     print(type(inst))  # the exception instance
                                #     # print(inst.args)  # arguments stored in .args
                                #     print(inst)
                                # finished_threads_no = thread_finish_q.qsize()
                                # print("finished_threads_no: " + str(finished_threads_no))
                                # for m in range(finished_threads_no):
                                #     next_state_q.put(current_state)
                                #     thread_id = int(thread_finish_q.get())
                                #     try:
                                #         sch_data.write(thread_id + 1, 7, np.array_str(current_state))
                                #         wb.save('Episodic_Data3.csv')
                                #     except Exception as inst:
                                #         print(type(inst))  # the exception instance
                                #         # print(inst.args)  # arguments stored in .args
                                #         print(inst)

                                while state_q.qsize() != 0 and reward_q.qsize() != 0 and action_q.qsize() != 0 and next_state_q.qsize() != 0 and done_q.qsize() != 0:
                                    print("Appending to buffer at episode end")
                                    # memory.append(
                                    #     (state_q.get(), action_q.get(), reward_q.get(), next_state_q.get(),
                                    #      done_q.get()))

                                    n_step_buffer.append(
                                        (state_q.get(), action_q.get(), reward_q.get(), next_state_q.get(),
                                         done_q.get()))
                                    print("n-step buffer length: ", len(n_step_buffer))

                                    if len(n_step_buffer) == n_step:  # fill the n-step buffer for the first translation
                                        # add a multi step transition
                                        reward, next_obs, done = get_n_step_info(n_step_buffer, gamma_nstep)
                                        obs, action = n_step_buffer[0][:2]

                                        memory.append((obs, action, reward, next_obs, done))
                                    # print("n_step_buffer length at episode end: ", len(n_step_buffer))
                                    # threadID_q.get()

                                    # print("n_step_buffer at episode end %s: " % (str(n_step_buffer)))

                                    # if len(n_step_buffer) == n_step:  # fill the n-step buffer for the first translation
                                    #     # add a multi step transition
                                    #     reward, next_obs, done = get_n_step_info(n_step_buffer, gamma_nstep)
                                    #     obs, action = n_step_buffer[0][:2]
                                    #
                                    #     # print("Appending to memory at episode end")
                                    #     memory.append((obs, action, reward, next_obs, done))

                                    print("memory size at episode end: ", len(memory))
                                    # print("memory at episode end %s: " % (str(memory)))

                                    # if threadID_q.qsize() != 0:
                                    #     for key in reward_dict:
                                    #         if key == threadID_q.queue[0]:
                                    #             print(
                                    #                 "AT END OF RUN Adding reward of thread %s to queue after memory update" % str(
                                    #                     key))
                                    #             reward_q.put(reward_dict[key])
                                    #             reward_dict.pop(key)
                                    #             # print("Now thread dict %s " % str(reward_dict))
                                    #             break

                                global epsilon
                                if len(memory) > train_start:
                                    if epsilon > epsilon_min:
                                        epsilon *= epsilon_decay
                                        print("changed Epsilon: " + str(epsilon))

                                try:
                                    sch_data.write(z + 1, 0, datetime.now())
                                    sch_data.write(z + 1, 1, e)
                                    sch_data.write(z + 1, 2, z)
                                    sch_data.write(z + 1, 3, event['object'].metadata.name)
                                    # sch_data.write(z + 1, 4, np.array_str(current_state))
                                    sch_data.write(z + 1, 5, 'None')
                                    sch_data.write(z, 7, np.array_str(next_state))
                                    sch_data.write(z, 8, done)
                                    sch_data.write(z + 1, 9, "Episode End")

                                    wb.save("episodic_data/Episodic_Data" + str(e) + ".csv")
                                    # print("Saved to Episodic data3")
                                except Exception as inst:
                                    print(type(inst))  # the exception instance
                                    # print(inst.args)  # arguments stored in .args
                                    # print(inst)

                                write_log = True
                                self.replay(write_log)
                                # if z % self.update_rate == 0:
                                #     self.update_target_network()

                                break
                            else:

                                current_state = self.get_current_state(fn_to_schedule, z + 1, e, sch_data,
                                                                       wb)
                                current_state = np.reshape(current_state, [1, self.state_size])
                                # print("Current state: ", current_state)

                                # available_act = self.available_actions(fn_to_schedule)
                                # print("Available nodes for scheduling: \n")
                                # for nodes in available_act:
                                #     print("Node: " + NODE_MAP_IP_NAME[nodes])

                                # finished_threads_no = thread_finish_q.qsize()
                                # print("finished_threads_no: " + str(finished_threads_no))
                                # for m in range(finished_threads_no):
                                #     next_state_q.put(current_state)
                                #     thread_id = int(thread_finish_q.get())
                                #     try:
                                #         sch_data.write(thread_id + 1, 7, np.array_str(current_state))
                                #         wb.save('Episodic_Data3.csv')
                                #     except Exception as inst:
                                #         print(type(inst))  # the exception instance
                                #         # print(inst.args)  # arguments stored in .args
                                #         print(inst)

                                # print("record action type in row: ", z + 1)

                                # if the fn has no incoming requests, do not include this pod creation as a data point

                                # *****************ADDDDDDDD***********************
                                if no_arrivals:
                                    res = self.scheduler(event['object'].metadata.name, "node-3")
                                    if not res:
                                        destroyed_pod_list.append(event['object'].metadata.name)
                                        print("Pod %s added to destroyed_pod_list and exiting" % (
                                            str(event['object'].metadata.name)))
                                    continue

                                # *******************************************************

                                global step_count_total
                                inference.write(step_count_total + 1, 0, e)
                                inference.write(step_count_total + 1, 1, z)
                                bef_time = time.time()
                                action, act_type = self.act(current_state, z + 1, e, sch_data, wb, fn_to_schedule)
                                inf_time = time.time() - bef_time
                                inference.write(step_count_total + 1, 2, inf_time)
                                wb_inf.save("results/Inference_time.csv")

                                # print("here")

                                res = self.scheduler(event['object'].metadata.name, NODE_MAP_INDEX_NAME[action])

                                if not res:
                                    destroyed_pod_list.append(event['object'].metadata.name)
                                    print("Pod %s added to destroyed_pod_list and exiting" % (
                                        str(event['object'].metadata.name)))
                                    # print("destroyed_pod_list: %s" % (str(destroyed_pod_list)))
                                    continue

                                pod_schd = datetime.now()
                                print("%s : Pod %s sent for scheduling to node %s: " % (
                                    pod_schd, event['object'].metadata.name, NODE_MAP_INDEX_NAME[action]))

                                step_count_total += 1
                                print("Step count: " + str(step_count_total))

                                # Active node penalty
                                active_node_penalty = 0
                                old_node = False
                                # current_active_nodes = get_active_nodes(event['object'].metadata.name)
                                # print("%s : CUrrent active nodes: %s" % (datetime.now(), str(current_active_nodes)))

                                for node_name in active_node_list:
                                    if node_name == NODE_MAP_INDEX_NAME[action]:
                                        old_node = True
                                        break

                                if not old_node:
                                    active_node_penalty = 15
                                else:
                                    active_node_penalty = -15

                                # app_latency = get_fn_latency_ratio(NODE_MAP_INDEX_NAME[action])
                                # print("%s : Initial App latency in node %s is %s: " % (datetime.now(), NODE_MAP_INDEX_NAME[action], str(app_latency)))

                                ##**
                                # pending_pod_list.append(event['object'].metadata.name)
                                ##**

                                # print("Response: " + str(res))

                                # end = datetime.now()
                                # print("% End Timestamp: ", end)
                                # print("Selected node for scheduling pod " + event['object'].metadata.name + ": ")
                                # print("Node: " + NODE_MAP_INDEX_NAME[action])

                                # print("Pod scheduled in node: " + NODE_MAP_INDEX_NAME[action])

                                if z != 0:
                                    next_state_q.put(current_state)
                                    next_state = current_state
                                    done_q.put(done)
                                    try:
                                        sch_data.write(z, 7, np.array_str(next_state))
                                        sch_data.write(z, 8, done)
                                    except Exception as inst:
                                        print(type(inst))  # the exception instance
                                        # print(inst.args)  # arguments stored in .args
                                        # print(inst)

                                state_q.put(current_state)
                                action_q.put(action)
                                # threadID_q.put(z)

                                thread_name = str(z)
                                util_penalty = 0
                                if cpu_util_dict[NODE_MAP_INDEX_NAME[action]] > 70:
                                    util_penalty = 10

                                set_reward_new(str(z), NODE_MAP_NAME_IP[NODE_MAP_INDEX_NAME[action]],
                                               z + 1,
                                               event['object'].metadata.name, e, sch_data, wb,
                                               active_node_penalty, step_count_total, thread_name, util_penalty)

                                # thread_name = str(z)
                                # thread_name = myThread(z, thread_name, NODE_MAP_NAME_IP[NODE_MAP_INDEX_NAME[action]],
                                #                        z + 1,
                                #                        event['object'].metadata.name, e, sch_data, wb,
                                #                        active_node_penalty, step_count_total
                                #                        )
                                # thread_name.start()

                                # if len(memory) > train_start:
                                #     if step_count_total % 20 == 0:
                                #         if epsilon > epsilon_min:
                                #             epsilon *= epsilon_decay
                                #             print("changed Epsilon: " + str(epsilon))

                                # if done:
                                #     print("episode: {}/{}, e: {:.2}".format(e, self.EPISODES, self.epsilon))
                                # if i == 500:
                                #     print("Saving trained model as cartpole-dqn.h5")
                                #     self.save("cartpole-dqn.h5")
                                #     return

                                sch_data.write(z + 1, 0, pod_schd)
                                sch_data.write(z + 1, 1, e)
                                sch_data.write(z + 1, 2, z)
                                sch_data.write(z + 1, 3, event['object'].metadata.name)
                                sch_data.write(z + 1, 4, np.array_str(current_state))
                                sch_data.write(z + 1, 5, NODE_MAP_INDEX_NAME[action])
                                sch_data.write(z + 1, 9, act_type)
                                # sch_data.write(z + 1, 8, done)

                                wb.save("episodic_data/Episodic_Data" + str(e) + ".csv")
                                # print("Saved to Episodic data3")

                                # if step_count_total % 50 == 0:
                                #     # agent.tensorboard.step += 1
                                #     write_log = True

                                self.replay(write_log)
                                if step_count_total % self.update_rate == 0 and len(memory) > train_start:
                                    print("Updating target network")
                                    self.update_target_network()

                                z = z + 1
                                # print("z value is now: ", z)

                        except client.ApiException as ex:
                            print(json.loads(ex.body)['message'])

            except Exception as ex:
                print("An exception has occurred in Kubernetes watch API :%s" % (str(ex)))

            print(
                "episode: {}/{}, e: {:.2}, episodic reward: {}".format(e, self.EPISODES, float(epsilon),
                                                                       episodic_reward))

            try:
                ep_data.write(e + 1, 0, datetime.now())
                ep_data.write(e + 1, 1, e)
                ep_data.write(e + 1, 2, epsilon)
                ep_data.write(e + 1, 3, episodic_reward)
                wb.save("episodic_data/Episodic_Data" + str(e) + ".csv")
                # print("Saved to Episodic data3")

            except Exception as inst:
                # print(type(inst))  # the exception instance
                print(inst.args)  # arguments stored in .args

            time.sleep(10)
            # print(inst)
            # *******************************************************Initial ENV
            m = 0
            n = 0
            try:
                w = watch.Watch()

                for event in w.stream(v1.list_namespaced_pod, "default"):
                    if event['object'].status.phase == "Pending" and event['object'].status.conditions == None and \
                            event[
                                'object'].spec.scheduler_name == scheduler_name:
                        try:
                            start = datetime.now()
                            print("% *************************Start of RESETTING Timestamp: ", start)
                            fn_to_schedule = event['object'].metadata.name.split('-')[0]
                            if fn_to_schedule == 'float':
                                n = 0
                            elif fn_to_schedule == 'primary':
                                n = 5
                            elif fn_to_schedule == 'load':
                                n = 0
                            elif fn_to_schedule == 'dd':
                                n = 0
                            elif fn_to_schedule == 'matrix':
                                n = 4
                            elif fn_to_schedule == 'imageprocess':
                                n = 1
                            elif fn_to_schedule == 'imageresize':
                                n = 1
                            elif fn_to_schedule == 'todoadditem':
                                n = 5
                            elif fn_to_schedule == 'tododelitem':
                                n = 9
                            elif fn_to_schedule == 'todogitem':
                                n = 9
                            elif fn_to_schedule == 'todolistitem':
                                n = 8
                            elif fn_to_schedule == 'todoupitem':
                                n = 6
                            elif fn_to_schedule == 'uploadimage':
                                n = 8
                            elif fn_to_schedule == 'uploadvideo':
                                n = 6
                            elif fn_to_schedule == 'video':
                                n = 1
                            elif fn_to_schedule == 'test':
                                training_end = True

                            res = self.scheduler(event['object'].metadata.name, NODE_MAP_INDEX_NAME[n])
                            print(res)
                            print("Selected node for scheduling pod " + event['object'].metadata.name + ": ")
                            print("Node: " + NODE_MAP_INDEX_NAME[n])
                            if training_end:
                                break
                            m += 1
                            if m > 14:
                                break

                        except client.ApiException as ex:
                            print(json.loads(ex.body)['message'])

            except Exception as ex:
                print("An exception has occurred in Kubernetes watch API :%s" % (str(ex)))

            # global training_end
            if training_end:
                break

        # *******************************************************************
        time_now = int(time.time())
        print("Saving trained model as Serverless_Scheduling %s.h5" % str(time_now))
        self.save(f"model/{MODEL_NAME}-{time_now}.h5")

        # # state = np.reshape(state, [1, self.state_size])
        # done = False
        # i = 0
        # while not done:
        #     self.env.render()
        #     action = self.act(state)
        #     next_state, reward, done, _ = self.env.step(action)
        #     # print(next_state)
        #     # print(reward)
        #     next_state = np.reshape(next_state, [1, self.state_size])
        #     if not done or i == self.env._max_episode_steps - 1:
        #         reward = reward
        #     else:
        #         reward = -100
        #     self.remember(state, action, reward, next_state, done)
        #     state = next_state
        #     i += 1
        #     if done:
        #         print("episode: {}/{}, score: {}, e: {:.2}".format(e, self.EPISODES, i, self.epsilon))
        #         if i == 500:
        #             print("Saving trained model as cartpole-dqn.h5")
        #             self.save("cartpole-dqn.h5")
        #             return
        #     self.replay()

    def test(self):
        self.load("model/Serverless_Scheduling-1661536194.h5")
        wb_inf = Workbook()
        inference = wb_inf.add_sheet('Inf_time')
        inference.write(0, 0, 'Episode')
        inference.write(0, 1, 'Event_No')
        inference.write(0, 2, 'Inf_time')
        print("Evaluation session started and sheet added")
        for e in range(self.EPISODES_TEST):
            global training_end
            global episodic_reward
            episodic_reward = 0
            global function_latency
            function_latency = 0
            global total_vm_time_diff
            total_vm_time_diff = 0
            global total_vm_cost_diff
            total_vm_cost_diff = 0
            global act_node_penalty
            act_node_penalty = 0
            global cpu_util_penalty
            cpu_util_penalty = 0
            global epsilon
            destroyed_pod_list = []
            global initial_lat
            initial_lat = 0
            wb = Workbook()
            sch_data = wb.add_sheet('Scheduled')
            sch_data.write(0, 0, 'Time')
            sch_data.write(0, 1, 'Episode')
            sch_data.write(0, 2, 'Event_No')
            sch_data.write(0, 3, 'Pod')
            sch_data.write(0, 4, 'State')
            sch_data.write(0, 5, 'Action')
            sch_data.write(0, 6, 'Reward')
            sch_data.write(0, 7, 'Next_state')
            sch_data.write(0, 8, 'Done')
            sch_data.write(0, 9, 'Random or NOT')
            sch_data.write(0, 10, 'Pod_created_time')
            sch_data.write(0, 11, 'Fn_latency')
            sch_data.write(0, 12, 'Fn_failure_rate')
            sch_data.write(0, 13, 'CPU_penalty')
            sch_data.write(0, 14, 'Active_node_no')
            sch_data.write(0, 15, 'VM_up_time_cost')
            sch_data.write(0, 16, 'VM_up_time')

            ep_data = wb.add_sheet('Episodes')
            ep_data.write(0, 0, 'Time')
            ep_data.write(0, 1, 'Episode')
            ep_data.write(0, 2, 'Epsilon')
            ep_data.write(0, 3, 'Ep_reward')
            ep_data.write(0, 4, 'Avg_nodes')

            reward_q.queue.clear()
            state_q.queue.clear()
            action_q.queue.clear()
            done_q.queue.clear()
            next_state_q.queue.clear()
            threadID_q.queue.clear()
            global reward_dict
            reward_dict = {}
            z = 0
            done = False
            start_thread = False

            # state = self.env.reset()
            event = 0

            try:
                w = watch.Watch()
                for event in w.stream(v1.list_namespaced_pod, "default"):

                    if event['object'].status.phase == "Pending" and event['object'].status.conditions == None and \
                            event[
                                'object'].spec.scheduler_name == scheduler_name:
                        write_log = False

                        # if z == 0:
                        #     global stop_an_thread
                        #     stop_an_thread = False
                        #     act_n_thread = "ep" + str(e)
                        #     active_node_thread = epThread_test(act_n_thread, ep_data, wb, e + 1, e)
                        #     active_node_thread.start()

                        if z == 0 and start_thread is False:
                            global stop_an_thread
                            print("starting vm info thread %s from main loop" % ("ep" + str(e)))
                            stop_an_thread = False
                            start_thread = True
                            act_n_thread = "ep" + str(e)
                            active_node_thread = epThread(act_n_thread, ep_data, wb, e + 1, e)
                            active_node_thread.start()

                        try:
                            start = datetime.now()
                            print("% *************************Start Timestamp: ", start)

                            pod_false = False
                            for dest_pod in destroyed_pod_list:
                                if event['object'].metadata.name == dest_pod:
                                    pod_false = True
                                    print("Pod is in the destroyed pod list so not scheduling it")
                                    break
                            if pod_false:
                                continue

                            fn_to_schedule = event['object'].metadata.name.split('-')[0]

                            if fn_to_schedule == "nginx":
                                # global stop_an_thread
                                stop_an_thread = True
                                print("NGINX pod")
                                # done = True
                                # done_q.put(done)
                                # reward_q.put(0)
                                current_state = np.zeros(77)
                                current_state = np.reshape(current_state, [1, self.state_size])
                                # state_q.put(current_state)
                                # next_state_q.put(current_state)
                                next_state = current_state
                                res = self.scheduler(event['object'].metadata.name, "node-3")

                                try:
                                    sch_data.write(z + 1, 0, datetime.now())
                                    sch_data.write(z + 1, 1, e)
                                    sch_data.write(z + 1, 2, z)
                                    sch_data.write(z + 1, 3, event['object'].metadata.name)
                                    # sch_data.write(z + 1, 4, np.array_str(current_state))
                                    sch_data.write(z + 1, 5, 'None')
                                    sch_data.write(z, 7, np.array_str(next_state))
                                    sch_data.write(z, 8, done)
                                    sch_data.write(z + 1, 9, "Episode End")

                                    wb.save("results/Evaluation_Data" + str(e) + ".csv")
                                    # print("Saved to Episodic data3")
                                except Exception as inst:
                                    print(type(inst))  # the exception instance

                                write_log = True
                                self.graphs(write_log)

                                break
                            else:

                                current_state = self.get_current_state(fn_to_schedule, z + 1, e, sch_data,
                                                                       wb)
                                current_state = np.reshape(current_state, [1, self.state_size])

                                # if the fn has no incoming requests, do not include this pod creation as a data point
                                if no_arrivals:
                                    res = self.scheduler(event['object'].metadata.name, "node-3")
                                    if not res:
                                        destroyed_pod_list.append(event['object'].metadata.name)
                                        print("Pod %s added to destroyed_pod_list and exiting" % (
                                            str(event['object'].metadata.name)))
                                    continue

                                global step_count_total
                                inference.write(step_count_total + 1, 0, e)
                                inference.write(step_count_total + 1, 1, z)
                                bef_time = time.time()
                                action = np.argmax(self.main_network.predict(current_state))
                                inf_time = time.time() - bef_time
                                inference.write(step_count_total + 1, 2, inf_time)
                                wb_inf.save("results/Inference_time.csv")

                                act_type = "Network"

                                res = self.scheduler(event['object'].metadata.name, NODE_MAP_INDEX_NAME[action])

                                if not res:
                                    destroyed_pod_list.append(event['object'].metadata.name)
                                    print("Pod %s added to destroyed_pod_list and exiting" % (
                                        str(event['object'].metadata.name)))
                                    # print("destroyed_pod_list: %s" % (str(destroyed_pod_list)))
                                    continue

                                pod_schd = datetime.now()
                                print("%s : Pod %s sent for scheduling to node %s: " % (
                                    pod_schd, event['object'].metadata.name, NODE_MAP_INDEX_NAME[action]))
                                step_count_total += 1
                                print("Step count: " + str(step_count_total))

                                # Active node penalty
                                active_node_penalty = 0
                                old_node = False
                                # current_active_nodes = get_active_nodes(event['object'].metadata.name)
                                # print("%s : CUrrent active nodes: %s" % (datetime.now(), str(current_active_nodes)))

                                for node_name in active_node_list:
                                    if node_name == NODE_MAP_INDEX_NAME[action]:
                                        old_node = True
                                        break

                                if not old_node:
                                    active_node_penalty = 15
                                else:
                                    active_node_penalty = -15

                                if z != 0:
                                    # next_state_q.put(current_state)
                                    next_state = current_state
                                    # done_q.put(done)
                                    try:
                                        sch_data.write(z, 7, np.array_str(next_state))
                                        sch_data.write(z, 8, done)
                                    except Exception as inst:
                                        print(type(inst))  # the exception instance
                                        # print(inst.args)  # arguments stored in .args
                                        # print(inst)

                                # state_q.put(current_state)
                                # action_q.put(action)
                                # threadID_q.put(z)

                                thread_name = str(z)
                                util_penalty = 0
                                if cpu_util_dict[NODE_MAP_INDEX_NAME[action]] > 70:
                                    util_penalty = 10

                                set_reward_test(str(z), NODE_MAP_NAME_IP[NODE_MAP_INDEX_NAME[action]],
                                                z + 1, event['object'].metadata.name,
                                                e, sch_data, wb,
                                                active_node_penalty, step_count_total, thread_name, util_penalty)

                                sch_data.write(z + 1, 0, pod_schd)
                                sch_data.write(z + 1, 1, e)
                                sch_data.write(z + 1, 2, z)
                                sch_data.write(z + 1, 3, event['object'].metadata.name)
                                sch_data.write(z + 1, 4, np.array_str(current_state))
                                sch_data.write(z + 1, 5, NODE_MAP_INDEX_NAME[action])
                                sch_data.write(z + 1, 9, act_type)


                                # sch_data.write(z + 1, 8, done)

                                wb.save("results/Evaluation_Data" + str(e) + ".csv")
                                # print("Saved to Episodic data3")

                                # if step_count_total % 50 == 0:
                                #     # agent.tensorboard.step += 1
                                #     write_log = True

                                # self.replay(write_log)
                                # if step_count_total % self.update_rate == 0 and len(memory) > train_start:
                                #     print("Updating target network")
                                #     self.update_target_network()

                                z = z + 1
                                # print("z value is now: ", z)

                        except client.ApiException as e:
                            print(json.loads(e.body)['message'])

            except Exception as ex:
                print("An exception has occurred in Kubernetes watch API :%s" % (str(ex)))

            print(
                "episode: {}/{}, e: {:.2}, episodic reward: {}".format(e, self.EPISODES, float(epsilon),
                                                                       episodic_reward))

            try:
                ep_data.write(e + 1, 0, datetime.now())
                ep_data.write(e + 1, 1, e)
                ep_data.write(e + 1, 2, epsilon)
                ep_data.write(e + 1, 3, episodic_reward)
                wb.save("results/Evaluation_Data" + str(e) + ".csv")
                # print("Saved to Episodic data3")

            except Exception as inst:
                # print(type(inst))  # the exception instance
                print(inst.args)  # arguments stored in .args

            if training_end:
                break

            time.sleep(10)
            # print(inst)
            # *******************************************************Initial ENV
            m = 0
            n = 0
            w = watch.Watch()

            for event in w.stream(v1.list_namespaced_pod, "default"):
                if event['object'].status.phase == "Pending" and event['object'].status.conditions == None and event[
                    'object'].spec.scheduler_name == scheduler_name:
                    try:
                        start = datetime.now()
                        print("% *************************Start of RESETTING Timestamp: ", start)
                        fn_to_schedule = event['object'].metadata.name.split('-')[0]
                        if fn_to_schedule == 'float':
                            n = 0
                        elif fn_to_schedule == 'primary':
                            n = 5
                        elif fn_to_schedule == 'load':
                            n = 0
                        elif fn_to_schedule == 'dd':
                            n = 0
                        elif fn_to_schedule == 'matrix':
                            n = 4
                        elif fn_to_schedule == 'imageprocess':
                            n = 1
                        elif fn_to_schedule == 'imageresize':
                            n = 1
                        elif fn_to_schedule == 'todoadditem':
                            n = 5
                        elif fn_to_schedule == 'tododelitem':
                            n = 9
                        elif fn_to_schedule == 'todogitem':
                            n = 9
                        elif fn_to_schedule == 'todolistitem':
                            n = 8
                        elif fn_to_schedule == 'todoupitem':
                            n = 6
                        elif fn_to_schedule == 'uploadimage':
                            n = 8
                        elif fn_to_schedule == 'uploadvideo':
                            n = 6
                        elif fn_to_schedule == 'video':
                            n = 1

                        res = self.scheduler(event['object'].metadata.name, NODE_MAP_INDEX_NAME[n])
                        print(res)
                        print("Selected node for scheduling pod " + event['object'].metadata.name + ": ")
                        print("Node: " + NODE_MAP_INDEX_NAME[n])
                        m += 1
                        if m > 5:
                            break

                    except client.ApiException as ex:
                        print(json.loads(ex.body)['message'])

        # *******************************************************************
        time_now = int(time.time())
        print("Ending Model Evaluation %s" % str(time_now))
        # self.save(f"model/{MODEL_NAME}-{time_now}.h5")


#
if __name__ == "__main__":
    config.load_kube_config()
    v1 = client.CoreV1Api()

    agent = DQNAgent()
    agent.tensorboard.step = 0

    # agent.run()
    agent.test()
